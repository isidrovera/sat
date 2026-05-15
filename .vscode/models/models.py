# -*- coding: utf-8 -*-

from odoo import _, models, fields, api
from odoo.exceptions import ValidationError
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta
import xml.etree.ElementTree as ET
import logging
_logger = logging.getLogger(__name__)
import xlwt
from io import BytesIO
import base64
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import tempfile
import os
import io
import qrcode
import requests
class SatSat(models.Model):
    _name = 'sat.sat'
    _description = 'Registro de maquina'
    
    _inherit = ['mail.thread', 'mail.activity.mixin', 'product.catalog.mixin']

    name = fields.Many2one('modelo.maquina', string='Modelo', required=True, tracking=True, 
     )
       

    @api.constrains('serie_id')
    def unique_field_serie_id(self):
        for item in self:
            items = self.search(
                [('serie_id', '=', item.serie_id), ('id', '!=', item.id)])
            if len(items) >= 1:
                raise ValidationError(_("El número de serie debe ser único."))


    reparaciones_count = fields.Integer(string='Reparaciones', compute='_compute_reparaciones_count')

    reparaciones_ids = fields.One2many('reparaciones.reparaciones', 'maquina_id')

    def _compute_reparaciones_count(self):
        for record in self:
            record.reparaciones_count = len(self.reparaciones_ids)

    def view_reparaciones_ids(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": "Reparaciones",
            "view_mode": "list,form",
            "res_model": "reparaciones.reparaciones",
            "domain": [("maquina_id", "=", self.id)],
            "context": "{'create': True}",
        }

    def action_partes(self):
        self.estado_ventas_id = "de_partes"

        self.disponibilidad_id = "no_disponible"

    def action_problemas(self):
        self.estado_ventas_id = "con_problemas"
        self.disponibilidad_id = "no_disponible"

    def action_sin_revisar(self):
        self.estado_ventas_id = "sin_revisar"

    def action_finalizado(self):
        self.estado_ventas_id = "finalizado"

    def action_en_revision(self):
        self.estado_ventas_id = "en_revision"

    def action_entregada(self):
        self.estado_ventas_id = "entregada"

    def action_disponible(self):
        self.disponibilidad_id = 'disponible'

    @api.model
    def _default_currency_id(self):
        value = self.env['res.currency'].search(
            [('name', '=', 'USD')], limit=1)
        return value and value.id or False
    currency_id = fields.Many2one('res.currency', string='Currency', default=_default_currency_id)
    trabajadores_id = fields.Many2one('hr.employee', string='Trabajadores', tracking=True,  default=1)

    def obtener_estado_ventas_display(self, estado_ventas_id):
        """Obtener el texto legible para el estado de ventas"""
        field = self._fields['estado_ventas_id']
        return dict(field.selection).get(estado_ventas_id)
    # Campos existentes en tu modelo sat.sat
    
    # Método para obtener registros filtrados
    def get_filtered_records(self):
        domain = [('disponibilidad_id', '=', 'no_disponible'), ('estado_ventas_id', '!=', 'entregada')]
        records = self.search(domain)
        return records
    
    # Método para enviar registros filtrados por correo
    def send_records_by_email(self):
        records = self.get_filtered_records()
        table_html = self.generate_html_table(records)
        excel_file = self.generate_excel_file(records)  # Nuevo método para generar el archivo Excel
        self.send_email(table_html, excel_file)  # Se pasa el archivo Excel como parámetro
        
    def generate_excel_file(self, records):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        worksheet = workbook.active

        # Agregar encabezados
        headers = ['Importación', 'Proveedor', 'Marca', 'Nombre', 'Serie', 'Contómetro', 'Descripción','Estado']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f'{col_letter}1'] = header

        # Agregar datos de los registros
        for row_num, record in enumerate(records, 2):
            worksheet[f'A{row_num}'] = record.importacion
            worksheet[f'B{row_num}'] = record.proveedor_id.name
            worksheet[f'C{row_num}'] = record.marca
            worksheet[f'D{row_num}'] = record.name.name
            worksheet[f'E{row_num}'] = record.serie_id
            worksheet[f'F{row_num}'] = record.contometro
            worksheet[f'G{row_num}'] = record.descripcion
            worksheet[f'H{row_num}'] = self.obtener_estado_ventas_display(record.estado_ventas_id)

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    
    # Método para generar tabla HTML
    def generate_html_table(self, records):
        table_html = '<table style="border-collapse: collapse; width: 100%;">'
        table_html += '<tr>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Importación</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Proveedor</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Marca</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Nombre</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Serie</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Contómetro</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Descripción</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Estado</th>'
        table_html += '</tr>'

        for record in records:
            table_html += '<tr>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.importacion}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.proveedor_id.name}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.marca}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.name.name}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.serie_id}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.contometro}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.descripcion}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{self.obtener_estado_ventas_display(record.estado_ventas_id)}</td>'
            table_html += '</tr>'

        table_html += '</table>'
        return table_html

    # Método para enviar correo
    def send_email(self, table_html, excel_file):
        Attachment = self.env['ir.attachment']
        data = {
            'name': 'Maquinas con problemas.xlsx',
            'datas': base64.b64encode(excel_file.getvalue()),
            'res_model': 'sat.sat',
            'res_id': self.id,
        }
        attachment = Attachment.create(data)

        template = self.env.ref('sat.email_maquinas_id')

        # Elimina los adjuntos existentes
        template.attachment_ids.unlink()

        full_html = template.body_html
        # Elimina cualquier tabla existente en el contenido del correo
        full_html = re.sub(r'<table.*?</table>', '', full_html, flags=re.DOTALL)
        # Agrega la nueva tabla al contenido del correo
        full_html += table_html

        template.write({
            'body_html': full_html,
            'attachment_ids': [(4, attachment.id)]
        })
        template.send_mail(self.id, force_send=True)



    
    @api.model
    def execute_task(self):
        weekday = datetime.today().weekday()  # 0 is Monday, 6 is Sunday
        if 0 <= weekday < 5:  # Only run on weekdays
            records = self.get_filtered_records()
            table_html = self.generate_html_table(records)
            excel_file = self.generate_excel_file(records)
            self.send_email(table_html, excel_file)
        else:
            _logger.info("Skipping execution because it's a weekend.")

    nextcall = fields.Datetime('Next Call')

    def get_next_call(self):
        today = datetime.today()
        days_ahead = 7 - today.weekday() if today.weekday() < 5 else 2
        next_date = today + timedelta(days=days_ahead)
        return next_date.strftime('%Y-%m-%d') + ' 13:00:00'



    def action_whatsap(self):
        msg = "Cliente: %s" % (self.cliente_id.name)
        msg1 = " Modelo: %s" % (self.name.name)
        msg2 = " Serie: %s" % (self.serie_id)
        msg3 = " Importación: %s" % (self.importacion)
        msg4 = " Proveedor: %s" % (self.proveedor_id.name)
        msg5 = " Marca: %s" % (self.marca)
        msg6 = " Ubicación: %s" % (self.ubicacion_id)
        msg7 = " Estado: %s" % (self.obtener_estado_ventas_display(self.estado_ventas_id))

        # msg2 = (f'{msg}{msg1}')

        whatsapp_iu_url = 'https://api.whatsapp.com/send?phone=%s&text=%s' % (
            self.trabajadores_id.mobile_phone, (f'{msg3}%0A{msg4}%0A{msg}%0A{msg5}%0A{msg1}%0A{msg2}%0A{msg6}%0A{msg7}'))
        return {
            'type': 'ir.actions.act_url',
                    'target': 'new',
                    'url': whatsapp_iu_url
        }


    reparacion_id = fields.Many2one('reparaciones.reparaciones',string='Reparacion', )  

    fecha_separacion = fields.Date(string="Fecha de separado")

    serie_id = fields.Char(string='Serie', tracking=True, required=True )

    estado_ventas_id = fields.Selection([('sin_revisar', 'Sin revisar'),('para_revision', 'Para revision'),('asignado','Asignado'),('en_revision', 'En revisión'), ('finalizado', 'Finalizado'), ('con_problemas', 'Con problemas'), ('de_partes', 'De partes'), ('entregada', 'Entregada')],
                                        string='Estado de revisión',
                                        default='sin_revisar', tracking=True
                                        )
    #_sql_constraints = [("unique_serie_id", "unique (serie_id)",
                        # "El numero de serie que intenta agregar ya existe")]

    tipo_revision = fields.Selection([('paso_papel', 'Paso papel'), (
        'regular', 'Regular'), ('cliente_final', 'Cliente final')], tracking=True)
    prioridad = fields.Selection(
        [('bajaubicacion', 'Baja'), ('alta', 'Alta')], tracking=True)
    disponibilidad_id = fields.Selection([('disponible', 'Disponible'),
                                         ('separada', 'Separada'),
                                         ('no_disponible', 'No disponible')],
                                         default='disponible', tracking=True,
                                         readonly=True, compute='_compute_disponibilidad_id', store=True
                                         )
    ubicacion_id = fields.Selection([('primer_piso', 'Primer piso'), ('tercer_piso', 'Tercer piso'), ('segundo_local', 'Segundo local'), ('covida', 'Covida')],
                                    default='primer_piso', tracking=True,
                                    )
    importacion = fields.Char('Importación', required=True, tracking=True)
    invoice = fields.Char('Invoice', required=True, tracking=True)

    precio_compra = fields.Float('Precio de compra', tracking=True)
    proveedor_id = fields.Many2one('res.partner', string='Proveedor', tracking=True, required=True
                                   )
    cliente_id = fields.Many2one('res.partner', string='Cliente', tracking=True)
    cliente_nombre = fields.Char(related='cliente_id.name', string='Cliente', size=10)
    asesora_id = fields.Char(related='cliente_id.asesora_id.name', readonly=True, store=True,  string='Asesora de ventas'
                             )
    
    factura_venta = fields.Char('Factura N°', tracking=True)
    fecha_entrega = fields.Date('Fecha de entrega', tracking=True)
    activador = fields.Selection([('si', 'Si'), ('no', 'No')],
                                 string=' ',
                                 default='no', tracking=True
                                 )
    descripcion = fields.Text(string='Descripción', tracking=True)
    @api.onchange('descripcion')
    def _onchange_descripcion(self):
        _logger.debug('Onchange Description for Record IDs: %s', self.ids)
        # Verifica cada registro en el conjunto
        for record in self:
            if record.descripcion:
                record.activador = 'si'
            else:
                record.activador = 'no'
        # Registra el estado final después de evaluar la descripción
        _logger.debug('Activador set to: %s', ', '.join([rec.activador for rec in self]))

    contometro = fields.Char(string='Contometro', required=True, tracking=True,)
    marca = fields.Char(string='Marca', related='name.marca_id.name', readonly=True, store=True, tracking=True
                        )
    precio_venta = fields.Float(string='Precio de venta', related='name.precio_venta', readonly=True, tracking=True)
    tipo_id = fields.Selection([('color', 'Color'), ('monocromatica', 'Monocromatica')],
                               string='Tipo de maquina', related='name.tipo_id', readonly=True, tracking=True)
    
    tipo_maquina = fields.Char(related='name.tipo_maquina_id.name', readonly=True, string='Tipo de maquina', tracking=True)

    

    
    @api.depends('cliente_id', 'estado_ventas_id')
    def _compute_disponibilidad_id(self):
        for record in self:
            _logger.debug('Computing Disponibilidad for Record ID: %s', record.id)
            # Incluyendo 'para_revision' en la lógica de disponibilidad
            if record.estado_ventas_id in ['sin_revisar', 'en_revision', 'finalizado', 'para_revision'] and record.cliente_id:
                _logger.debug('Setting disponibilidad_id to separada for Record ID: %s', record.id)
                record.disponibilidad_id = 'separada'
                record.fecha_separacion = fields.Date.today()
            elif record.estado_ventas_id in ['sin_revisar', 'en_revision', 'finalizado', 'para_revision'] and not record.cliente_id:
                _logger.debug('Setting disponibilidad_id to disponible for Record ID: %s', record.id)
                record.disponibilidad_id = 'disponible'
                record.fecha_separacion = False
            else:
                _logger.debug('Setting disponibilidad_id to no disponible for Record ID: %s', record.id)
                record.disponibilidad_id = 'no_disponible'
                record.fecha_separacion = False
            _logger.debug('Disponibilidad ID updated to %s for Record ID: %s', record.disponibilidad_id, record.id)



    @api.onchange('factura_venta')
    def onchange_factura_venta(self):
        if self.factura_venta:
            self.estado_ventas_id = 'entregada'
            
            
    
    def generate_record_url(self, record):
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        action_id = self.env.ref('sat.action_window').id
        menu_id = self.env.ref('sat.stock_maquinas').id
        url = "{}/web#id={}&view_type=form&model=sat.sat&action={}&menu_id={}".format(base_url, record.id, action_id, menu_id)
        return url
    qr_image = fields.Binary("QR Image", compute="generate_qr_code", attachment=True, store=True)


    @api.depends('estado_ventas_id')  # Suponiendo que quieras codificar un campo específico, reemplaza 'nombre_del_campo_a_codificar' con el campo relevante.
    def generate_qr_code(self):
        for record in self:
            # Generar la URL del registro
            url = self.generate_record_url(record)
            
            if not url:
                continue
            
            # Crear un objeto de código QR
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(url)
            qr.make(fit=True)
            
            # Generar la imagen del código QR
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Guardar la imagen en un buffer en memoria
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            
            # Codificar la imagen en base64
            img_base64 = base64.b64encode(buffer.read())
            
            # Guardar la imagen codificada en el campo qr_image
            record.qr_image = img_base64
    icono_rojo = fields.Html(compute='_compute_icono_rojo', string=' ')

    @api.depends('activador')  # Reemplaza con el campo real que afecta la condición
    def _compute_icono_rojo(self):
        for record in self:
            # Aquí va la lógica para determinar cuándo el ícono debe ser rojo
            if record.activador:  # Reemplaza con tu condición real
                record.icono_rojo = '<i class="fa fa-exclamation-circle icono-grande" style="color: red;"></i>'
            else:
                record.icono_rojo = False

    def action_primer(self):
        self.ubicacion_id = 'primer_piso'

    def action_tercero(self):
        self.ubicacion_id = 'tercer_piso'

    def action_segundo(self):
        self.ubicacion_id = 'segundo_local'

    def create_reparacion(self):
        self.ensure_one()  # Asegurarse de que se trabaja con un único registro.
        return {
            'name': 'Crear Reparación',
            'type': 'ir.actions.act_window',
            'res_model': 'reparaciones.reparaciones',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_maquina_id': self.id,
                
            },
        }

    fecha_para_revision = fields.Datetime(string="Fecha para Revisión", readonly=True, tracking=True)


    def get_isidro_partner_id(self):
        isidro_user = self.env['res.users'].search([('name', '=', 'Isidro Vera Polo')], limit=1)
        if isidro_user:
            return isidro_user.partner_id.id
        return False

    def write(self, vals):
        estados_permitidos_para_cambio = ['sin_revisar', 'para_revision']
        estados_problema = ['con_problemas', 'de_partes']
        estado_final_no_notificar = 'entregada'

        tipo_revision_modificado = 'tipo_revision' in vals
        prioridad_modificada = 'prioridad' in vals

        isidro_partner_id = self.get_isidro_partner_id()

        # 📌 Snapshot ANTES de escribir, para detectar cambios de modelo/tipo/contómetro
        cambios_previos = {}
        for record in self:
            cambios_previos[record.id] = {
                'modelo_anterior': record.name.name if record.name else '',
                'tipo_anterior': record.tipo_id,
                'contometro_anterior': record.contometro or '0',
            }

        for record in self:
            estado_actual = record.estado_ventas_id
            nuevo_estado = vals.get('estado_ventas_id', estado_actual)

            _logger.debug(
                f"Inicio de write para ID {record.id}. "
                f"Estado actual: {estado_actual}, Nuevo estado: {nuevo_estado}, Valores: {vals}"
            )

            # Primera parte: Manejo de tipo_revision y prioridad
            if estado_actual in estados_permitidos_para_cambio:
                if tipo_revision_modificado or prioridad_modificada:
                    if vals.get('tipo_revision') or vals.get('prioridad'):
                        vals['estado_ventas_id'] = 'para_revision'
                        # Convertir la hora UTC a hora de Perú al guardar
                        utc_now = datetime.utcnow()
                        peru_tz = pytz.timezone('America/Lima')
                        peru_dt = pytz.utc.localize(utc_now).astimezone(peru_tz)
                        vals['fecha_para_revision'] = peru_dt.astimezone(pytz.UTC).strftime('%Y-%m-%d %H:%M:%S')

                        _logger.info(
                            f"Estado cambiado a 'para_revision' para ID {record.id} "
                            f"por modificación en tipo_revision o prioridad."
                        )

                        if isidro_partner_id:
                            user_name = self.env.user.name
                            record_name = record.name.name
                            serie = record.serie_id
                            message = f"""Se ha colocado una nueva máquina para revisión.

Detalles del equipo:
- Nombre: {record_name}
- Serie: {serie}
- Fecha de registro: {peru_dt.strftime('%Y-%m-%d %H:%M:%S')} (hora Lima)

Modificado por: {user_name}"""
                            record.message_post(
                                body=message,
                                partner_ids=[isidro_partner_id],
                                subtype='mail.mt_comment',
                            )
                    else:
                        vals['estado_ventas_id'] = 'sin_revisar'
                        vals['fecha_para_revision'] = None
                        _logger.info(
                            f"Estado regresado a 'sin_revisar' para ID {record.id}."
                        )

            # Segunda parte: Manejo de estados de problema y notificaciones
            if 'estado_ventas_id' in vals:
                nuevo_estado = vals['estado_ventas_id']

                # Si cambia a estado de problema
                if nuevo_estado in estados_problema:
                    _logger.debug(
                        f"Cambiando a estado de problema para ID {record.id}. "
                        f"Ejecutando super().write()."
                    )
                    result = super(SatSat, self).write(vals)

                    try:
                        record.enviar_mensaje_problema_asesora()
                        _logger.info(
                            f"Notificación de problema enviada para ID {record.id}."
                        )
                    except Exception as e:
                        _logger.error(
                            f"Error al enviar notificaciones para ID {record.id}: {e}"
                        )

                    return result

                # Si cambia de un estado problemático a otro no problemático
                elif estado_actual in estados_problema and nuevo_estado not in estados_problema:
                    _logger.debug(
                        f"Saliendo de estado de problema para ID {record.id}. "
                        f"Limpiando descripción."
                    )
                    vals['descripcion'] = False
                    vals['activador'] = 'no'
                    message = _(
                        "Se limpió la descripción al cambiar el estado de '%s' a '%s'"
                    ) % (estado_actual, nuevo_estado)
                    record.message_post(body=message)

                # Nueva lógica: Enviar notificación de disponibilidad si aplica
                if estado_actual in estados_problema and nuevo_estado != estado_final_no_notificar:
                    _logger.debug(
                        f"Enviando notificación de disponibilidad para ID {record.id}."
                    )
                    try:
                        record.enviar_notificacion_disponibilidad()
                        _logger.info(
                            f"Notificación de disponibilidad enviada para ID {record.id}."
                        )
                    except Exception as e:
                        _logger.error(
                            f"Error al enviar notificación de disponibilidad para ID {record.id}: {e}"
                        )

        # Ejecutar la escritura final después de todas las validaciones y notificaciones
        _logger.debug(f"Finalizando write para registros {self.ids} con valores: {vals}")
        result = super(SatSat, self).write(vals)

        # 🔍 Después de escribir, revisar anomalías de modelo y contómetro
        for record in self:
            prev = cambios_previos.get(record.id) or {}
            modelo_anterior = prev.get('modelo_anterior', '')
            tipo_anterior = prev.get('tipo_anterior')
            contometro_anterior = prev.get('contometro_anterior', '0')

            modelo_nuevo = record.name.name if record.name else ''
            tipo_nuevo = record.tipo_id
            contometro_nuevo = record.contometro or '0'

            # 1) Cambios raros de modelo (velocidad / color)
            self._check_model_anomalies(
                record,
                modelo_anterior,
                modelo_nuevo,
                tipo_anterior,
                tipo_nuevo,
            )

            # 2) Saltos raros de contómetro
            self._check_counter_anomalies(
                record,
                contometro_anterior,
                contometro_nuevo,
            )

        return result
    
    def _check_model_anomalies(self, record, modelo_anterior, modelo_nuevo, tipo_anterior, tipo_nuevo):
        """
        Detecta casos como:
        - Canon 4525  -> 4525/4535 (cambio de velocidad dentro misma familia)
        - bizhub 364e -> bizhub C364e (cambio de mono a color)
        y avisa por chatter/correo (a Isidro).
        """
        # Si no cambió el modelo, no hacemos nada
        if not modelo_anterior or not modelo_nuevo or modelo_anterior == modelo_nuevo:
            return

        # Extraer último bloque numérico de cada modelo (núcleo de velocidad)
        def _get_core_digits(text):
            nums = re.findall(r'\d+', text or '')
            return nums[-1] if nums else None

        core_old = _get_core_digits(modelo_anterior)
        core_new = _get_core_digits(modelo_nuevo)

        posible_cambio_velocidad = False
        detalle_velocidad = ""

        if core_old and core_new and core_old != core_new:
            # Caso típico: 4525 -> 4535 (mismo largo y primeros dígitos iguales)
            try:
                if len(core_old) == len(core_new):
                    if len(core_old) == 4 and core_old[:2] == core_new[:2]:
                        posible_cambio_velocidad = True
                        detalle_velocidad = f"{core_old[-2:]} → {core_new[-2:]}"
                    else:
                        # fallback: cualquier cambio de núcleo con mismo largo
                        posible_cambio_velocidad = True
                        detalle_velocidad = f"{core_old} → {core_new}"
            except Exception:
                pass

        # Cambio de tipo (color/mono)
        cambio_tipo = False
        if tipo_anterior and tipo_nuevo and tipo_anterior != tipo_nuevo:
            cambio_tipo = True

        # Si no hay nada relevante, salir
        if not posible_cambio_velocidad and not cambio_tipo:
            return

        isidro_partner_id = record.get_isidro_partner_id()
        url = record.generate_record_url(record)

        lineas = [
            "Se detectó una actualización relevante del modelo (posiblemente por SNMP o edición manual):",
            f"• Modelo anterior: <b>{modelo_anterior}</b>",
            f"• Modelo nuevo: <b>{modelo_nuevo}</b>",
        ]

        if posible_cambio_velocidad:
            lineas.append(f"• Posible cambio de velocidad (núcleo): <b>{detalle_velocidad}</b>")

        if cambio_tipo:
            sel_tipo = dict(record._fields['tipo_id'].selection)
            txt_old = sel_tipo.get(tipo_anterior, tipo_anterior)
            txt_new = sel_tipo.get(tipo_nuevo, tipo_nuevo)
            lineas.append(f"• Cambio de tipo: <b>{txt_old}</b> → <b>{txt_new}</b>")

        lineas.append(f"• Equipo: <b>{record.name.name if record.name else ''}</b> / Serie: <b>{record.serie_id}</b>")
        lineas.append(f"• Enlace al equipo: {url}")

        body = "<br/>".join(lineas)

        record.message_post(
            body=body,
            subtype_xmlid='mail.mt_note',
            partner_ids=[isidro_partner_id] if isidro_partner_id else None,
        )

    def _check_counter_anomalies(self, record, contometro_anterior, contometro_nuevo):
        """
        Detecta variaciones sospechosas en el contómetro, por ejemplo:
        - 2,000  → 20,000  (x10)
        - 2,000  → 2,000,000 (muchos más dígitos)
        y NO molesta si es algo normal, como:
        - 40,000 → 42,000
        """

        # Limpiar a solo dígitos
        old_digits = re.sub(r'[^\d]', '', contometro_anterior or '') or '0'
        new_digits = re.sub(r'[^\d]', '', contometro_nuevo or '') or '0'

        try:
            old_val = int(old_digits)
            new_val = int(new_digits)
        except Exception:
            return

        # Si alguno es cero o el nuevo es menor, no analizamos aquí (ya tienes lógica SNMP aparte)
        if old_val <= 0 or new_val <= 0 or new_val <= old_val:
            return

        digit_diff = abs(len(str(old_val)) - len(str(new_val)))
        ratio = new_val / float(old_val) if old_val else 0.0

        # Reglas:
        # - muchos más dígitos (ej: 4 -> 7)
        # - o incremento >= x10 del valor anterior
        if digit_diff < 2 and ratio < 10.0:
            # incremento normal, no avisamos
            return

        incremento = new_val - old_val
        isidro_partner_id = record.get_isidro_partner_id()
        url = record.generate_record_url(record)

        lineas = [
            "⚠️ Se detectó una variación inusual en el contómetro:",
            f"• Valor anterior: <b>{old_val:,}</b>",
            f"• Valor nuevo: <b>{new_val:,}</b>",
            f"• Incremento: <b>{incremento:,}</b>",
            f"• Multiplicador aproximado: <b>x{ratio:.1f}</b>",
            f"• Equipo: <b>{record.name.name if record.name else ''}</b> / Serie: <b>{record.serie_id}</b>",
            f"• Enlace al equipo: {url}",
        ]

        body = "<br/>".join(lineas)

        record.message_post(
            body=body,
            subtype_xmlid='mail.mt_note',
            partner_ids=[isidro_partner_id] if isidro_partner_id else None,
        )


    @api.onchange('disponibilidad_id', 'ubicacion_id')
    def _onchange_disponibilidad_ubicacion(self):
        if self.disponibilidad_id == 'separada' and self.ubicacion_id in ['segundo_local', 'covida']:
            self.enviar_mensaje_transportistas()
            return self._notify_vendedora()

    def enviar_mensaje_transportistas(self):
        transportista_numeros = ['51924894872']
        mensaje = f"Estimado transportista,\n\nPor favor, traer la siguiente máquina:\n\nModelo: {self.name.name}\nSerie: {self.serie_id}\nUbicación actual: {self.ubicacion_id}."
        url = self.crear_url_cambio_ubicacion(self)

        mensaje += f"\n\nPara cambiar la ubicación a primer piso, haga clic en el siguiente enlace: 📍 {url}"

        _logger.debug(f"Enviando mensaje a transportistas: {mensaje}")

        for numero in transportista_numeros:
            self.enviar_mensaje_whatsapp(numero, mensaje)

    def enviar_mensaje_whatsapp(self, phone, message):
        """Envía un mensaje de WhatsApp a un número o grupo."""
        url = 'https://boot.andessolutioncopiers.com/api/send-message'
        data = {
            'to': phone,
            'message': message
        }
        headers = {
            'Content-Type': 'application/json',
            'x-api-key': 'wg_fc215093f007df7ff4a32c04c7d8170d11960583e3a1b43a695037f5a627d3e3'
        }
        
        try:
            response = requests.post(url, headers=headers, json=data, timeout=30)
            response.raise_for_status()
            
            # Verificar respuesta exitosa
            response_data = response.json()
            if response_data.get('success'):
                _logger.info(f"✅ Mensaje enviado exitosamente a {phone}")
                return True
            else:
                error_msg = response_data.get('error', 'Error desconocido')
                _logger.error(f"❌ Error en API al enviar a {phone}: {error_msg}")
                return False
                
        except requests.exceptions.Timeout:
            _logger.error(f"❌ Timeout al enviar mensaje de WhatsApp a {phone}")
            return False
        except requests.exceptions.RequestException as e:
            _logger.error(f"❌ Error al enviar mensaje de WhatsApp a {phone}: {e}")
            return False
        except Exception as e:
            _logger.error(f"❌ Error inesperado al enviar WhatsApp a {phone}: {e}")
            return False
    def crear_url_cambio_ubicacion(self, record):
        """Genera una URL única para el cambio de ubicación, incluyendo un token."""
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        token = base64.b64encode(os.urandom(24)).decode()  # Generar un token único
        record.write({'location_change_token': token})  # Almacenar el token en el registro
        return f"{base_url}/sat/change_location/{record.id}?token={token}"

    def _notify_vendedora(self):
        return {
            'warning': {
                'title': "Notificación",
                'message': "Estimada vendedora, ya se está notificando a transporte que traigan el equipo.",
                'type': 'notification'
            }
        }
    @api.model
    def cron_evaluador_diario(self):
        _logger.debug("Iniciando cron_evaluador_diario")
        self.evaluar_registros_diarios()

    def evaluar_registros_diarios(self):
        registros_primer_piso = self.search([('ubicacion_id', '=', 'primer_piso'), ('estado_ventas_id', '=', 'sin_revisar')])
        registros_tercer_piso = self.search([('ubicacion_id', '=', 'tercer_piso'), ('estado_ventas_id', '=', 'sin_revisar')])

        if not registros_primer_piso and not registros_tercer_piso:
            registros_a_traer = self.search([
                ('ubicacion_id', 'in', ['segundo_local', 'covida']),
                ('estado_ventas_id', '=', 'sin_revisar')
            ], limit=8)
            
            _logger.debug(f"Máquinas a traer: {registros_a_traer}")
            
            if registros_a_traer:
                transportista_numeros = ['51924894872']
                for registro in registros_a_traer:
                    mensaje = f"Estimado transportista,\n\nPor favor, traer la siguiente máquina:\n\nModelo: {registro.name.name}\nSerie: {registro.serie_id}\nUbicación actual: {registro.ubicacion_id}."
                    url = self.crear_url_cambio_ubicacion(registro)

                    mensaje += f"\n\nPara cambiar la ubicación a primer piso, haga clic en el siguiente enlace: 📍 {url}"

                    _logger.debug(f"Enviando mensaje para la máquina {registro.name.name} con serie {registro.serie_id}")

                    for numero in transportista_numeros:
                        self.enviar_mensaje_whatsapp(numero, mensaje)
                    
                    _logger.info(f"Mensaje enviado para la máquina {registro.name.name} con serie {registro.serie_id}")

    def action_crear_reparaciones(self):
        """ Crea una reparación para cada registro en el modelo 'sat.sat'. """
        # Buscar todos los registros en el modelo 'sat.sat'
        sat_records = self.search([])  # Esto obtiene todos los registros de la tabla 'sat.sat'
        
        reparacion_model = self.env['reparaciones.reparaciones']
        for record in sat_records:
            # Crear la reparación para cada registro
            reparacion_model.create({
                'maquina_id': record.id,  # Relaciona la reparación con el registro actual de 'sat.sat'
                  # Verifica si este campo existe en el modelo
                # Puedes agregar otros campos aquí según los necesarios en el modelo 'reparaciones.reparaciones'
            })
            _logger.info(f"Reparación creada para la máquina {record.name.name} con serie {record.serie_id}")

        # Mostrar una notificación de éxito
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Reparaciones creadas'),
                'message': _('Se han creado las reparaciones para todas las máquinas.'),
                'type': 'success',
                'sticky': False,
            }
        }

    def abrir_cambio_estado(self):
            # Aquí puedes abrir un wizard para cambiar el estado
            return {
                'type': 'ir.actions.act_window',
                'name': 'Cambiar Estado',
                'res_model': 'sat.cambio.estado.wizard',
                'view_mode': 'form',
                'target': 'new',
            }



    total_maquinas = fields.Integer(compute='_compute_maquinas')
    maquinas_disponibles = fields.Integer(compute='_compute_maquinas')
    maquinas_separadas = fields.Integer(compute='_compute_maquinas')
    maquinas_no_disponibles = fields.Integer(compute='_compute_maquinas')

    @api.depends('disponibilidad_id')
    def _compute_maquinas(self):
        for record in self:
            record.total_maquinas = self.search_count([])
            record.maquinas_disponibles = self.search_count([('disponibilidad_id', '=', 'disponible')])
            record.maquinas_separadas = self.search_count([('disponibilidad_id', '=', 'separada')])
            record.maquinas_no_disponibles = self.search_count([('disponibilidad_id', '=', 'no_disponible')])