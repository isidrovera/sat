# -*- coding: utf-8 -*-

from odoo import _, models, fields, api
from odoo.exceptions import ValidationError
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta
import xml.etree.ElementTree as ET
import logging
_logger = logging.getLogger(__name__)
import xlwt
from io import BytesIO
import base64
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import tempfile
import os
import io
import pytz
import qrcode
import requests
class SatSat(models.Model):
    _name = 'sat.sat'
    _description = 'Registro de maquina'
    
    _inherit = ['mail.thread', 'mail.activity.mixin', 'product.catalog.mixin']

    name = fields.Many2one('modelo.maquina', string='Modelo', required=True, tracking=True, 
     )
       

    @api.constrains('serie_id')
    def unique_field_serie_id(self):
        for item in self:
            items = self.search(
                [('serie_id', '=', item.serie_id), ('id', '!=', item.id)])
            if len(items) >= 1:
                raise ValidationError(_("El número de serie debe ser único."))

    @api.model
    def create(self, vals):
        """
        Al crear un nuevo registro, copia automáticamente el contómetro 
        al campo contometro_proveedor (valor inicial del proveedor).
        """
        # Si viene contómetro pero no contometro_proveedor, copiar automáticamente
        if vals.get('contometro') and not vals.get('contometro_proveedor'):
            vals['contometro_proveedor'] = vals['contometro']
            _logger.info(
                "[CREATE] Copiando contómetro → contometro_proveedor: %s",
                vals['contometro']
            )
        
        # Crear el registro normalmente
        record = super(SatSat, self).create(vals)
        
        return record



    
    reparaciones_count = fields.Integer(string='Reparaciones', compute='_compute_reparaciones_count')

    reparaciones_ids = fields.One2many('reparaciones.reparaciones', 'maquina_id')

    def _compute_reparaciones_count(self):
        for record in self:
            record.reparaciones_count = len(self.reparaciones_ids)

    def view_reparaciones_ids(self):
        self.ensure_one()
        # Buscar un registro de reparaciones que cumpla con el dominio
        reparacion = self.env["reparaciones.reparaciones"].search([("maquina_id", "=", self.id)], limit=1)
        
        return {
            "type": "ir.actions.act_window",
            "name": "Reparaciones",
            "view_mode": "form",
            "res_model": "reparaciones.reparaciones",
            "res_id": reparacion.id if reparacion else False,  # ID del registro a mostrar en el formulario
            "domain": [("maquina_id", "=", self.id)],
            "context": {'create': True},
        }

    autorizacion_cambio_digitos = fields.Boolean(string="Autorización de Modificación", default=False)
    

    @api.model
    def _default_currency_id(self):
        value = self.env['res.currency'].search(
            [('name', '=', 'USD')], limit=1)
        return value and value.id or False
    currency_id = fields.Many2one('res.currency', string='Currency', default=_default_currency_id)
    trabajadores_id = fields.Many2one('hr.employee', string='Trabajadores', tracking=True,  default=1)

    def obtener_estado_ventas_display(self, estado_ventas_id):
        """Obtener el texto legible para el estado de ventas"""
        field = self._fields['estado_ventas_id']
        return dict(field.selection).get(estado_ventas_id)
    # Campos existentes en tu modelo sat.sat
    
    # Método para obtener registros filtrados
    def get_filtered_records(self):
        domain = [('disponibilidad_id', '=', 'no_disponible'), ('estado_ventas_id', '!=', 'entregada')]
        records = self.search(domain)
        return records
    
    # Método para enviar registros filtrados por correo
    def send_records_by_email(self):
        records = self.get_filtered_records()
        table_html = self.generate_html_table(records)
        excel_file = self.generate_excel_file(records)  # Nuevo método para generar el archivo Excel
        self.send_email(table_html, excel_file)  # Se pasa el archivo Excel como parámetro
        
    def generate_excel_file(self, records):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        worksheet = workbook.active

        # Agregar encabezados
        headers = ['Importación', 'Proveedor', 'Marca', 'Nombre', 'Serie', 'Contómetro', 'Descripción','Estado']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f'{col_letter}1'] = header

        # Agregar datos de los registros
        for row_num, record in enumerate(records, 2):
            worksheet[f'A{row_num}'] = record.importacion
            worksheet[f'B{row_num}'] = record.proveedor_id.name
            worksheet[f'C{row_num}'] = record.marca
            worksheet[f'D{row_num}'] = record.name.name
            worksheet[f'E{row_num}'] = record.serie_id
            worksheet[f'F{row_num}'] = record.contometro
            worksheet[f'G{row_num}'] = record.descripcion
            worksheet[f'H{row_num}'] = self.obtener_estado_ventas_display(record.estado_ventas_id)

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    
    # Método para generar tabla HTML
    def generate_html_table(self, records):
        table_html = '<table style="border-collapse: collapse; width: 100%;">'
        table_html += '<tr>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Importación</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Proveedor</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Marca</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Nombre</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Serie</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Contómetro</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Descripción</th>'
        table_html += '<th style="border: 1px solid black; padding: 8px;">Estado</th>'
        table_html += '</tr>'

        for record in records:
            table_html += '<tr>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.importacion}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.proveedor_id.name}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.marca}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.name.name}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.serie_id}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.contometro}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{record.descripcion}</td>'
            table_html += f'<td style="border: 1px solid black; padding: 8px;">{self.obtener_estado_ventas_display(record.estado_ventas_id)}</td>'
            table_html += '</tr>'

        table_html += '</table>'
        return table_html

    # Método para enviar correo
    def send_email(self, table_html, excel_file):
        Attachment = self.env['ir.attachment']
        data = {
            'name': 'Maquinas con problemas.xlsx',
            'datas': base64.b64encode(excel_file.getvalue()),
            'res_model': 'sat.sat',
            'res_id': self.id,
        }
        attachment = Attachment.create(data)

        template = self.env.ref('sat.email_maquinas_id')

        # Elimina los adjuntos existentes
        template.attachment_ids.unlink()

        full_html = template.body_html
        # Elimina cualquier tabla existente en el contenido del correo
        full_html = re.sub(r'<table.*?</table>', '', full_html, flags=re.DOTALL)
        # Agrega la nueva tabla al contenido del correo
        full_html += table_html

        template.write({
            'body_html': full_html,
            'attachment_ids': [(4, attachment.id)]
        })
        template.send_mail(self.id, force_send=True)



    
    @api.model
    def execute_task(self):
        weekday = datetime.today().weekday()  # 0 is Monday, 6 is Sunday
        if 0 <= weekday < 5:  # Only run on weekdays
            records = self.get_filtered_records()
            table_html = self.generate_html_table(records)
            excel_file = self.generate_excel_file(records)
            self.send_email(table_html, excel_file)
        else:
            _logger.info("Skipping execution because it's a weekend.")

    nextcall = fields.Datetime('Next Call')

    def get_next_call(self):
        today = datetime.today()
        days_ahead = 7 - today.weekday() if today.weekday() < 5 else 2
        next_date = today + timedelta(days=days_ahead)
        return next_date.strftime('%Y-%m-%d') + ' 13:00:00'



    
    reparacion_id = fields.Many2one('reparaciones.reparaciones',string='Reparacion', )  

    fecha_separacion = fields.Date(string="Fecha de separado")

    serie_id = fields.Char(string='Serie', tracking=True, required=True )

    estado_ventas_id = fields.Selection([('sin_revisar', 'Sin revisar'),('para_revision', 'Para revision'),('en_revision', 'En revisión'), ('finalizado', 'Finalizado'), ('con_problemas', 'Con problemas'), ('de_partes', 'De partes'), ('entregada', 'Entregada')],
                                        string='Estado de revisión',
                                        default='sin_revisar', tracking=True
                                        )
    #_sql_constraints = [("unique_serie_id", "unique (serie_id)",
                        # "El numero de serie que intenta agregar ya existe")]

    tipo_revision = fields.Selection([('paso_papel', 'Paso papel'), (
        'regular', 'Regular'), ('cliente_final', 'Cliente final')], tracking=True)
    prioridad = fields.Selection(
        [('bajaubicacion', 'Baja'), ('alta', 'Alta')], tracking=True)
    disponibilidad_id = fields.Selection([('disponible', 'Disponible'),
                                         ('separada', 'Separada'),
                                         ('no_disponible', 'No disponible')],
                                         default='disponible', tracking=True,
                                         readonly=True, compute='_compute_disponibilidad_id', store=True
                                         )
    ubicacion_id = fields.Selection([('primer_piso', 'Primer piso'), ('tercer_piso', 'Tercer piso'), ('segundo_local', 'Segundo local'), ('covida', 'Covida')],
                                    default='primer_piso', tracking=True,
                                    )
    importacion = fields.Char('Importación', required=True, tracking=True)
    invoice = fields.Char('Invoice', required=True, tracking=True)

    precio_compra = fields.Float('Precio de compra', tracking=True)
    proveedor_id = fields.Many2one('res.partner', string='Proveedor', tracking=True, required=True
                                   )
    cliente_id = fields.Many2one('res.partner', string='Cliente', tracking=True)
    cliente_nombre = fields.Char(related='cliente_id.name', string='Cliente', size=10)
    asesora_id = fields.Char(related='cliente_id.asesora_id.name', readonly=True, store=True,  string='Asesora de ventas'
                             )
    
    factura_venta = fields.Char('Factura N°', tracking=True)
    fecha_entrega = fields.Date('Fecha de entrega', tracking=True)
    activador = fields.Selection([('si', 'Si'), ('no', 'No')],
                                 string=' ',
                                 default='no', tracking=True
                                 )
    descripcion = fields.Text(string='Descripción', tracking=True)
    check_ingreso = fields.Boolean(
        string='Ingreso registrado',
        help='Indica si el equipo ya fue ingresado/descargado mediante el scanner.'
    )

    ingreso_estado = fields.Selection([
        ('none', 'Sin check'),
        ('ok_no_obs', 'OK (sin observaciones)'),
        ('ok_obs', 'OK (con observaciones)'),
        ('rechazado', 'Rechazado'),
    ], string='Estado de ingreso', default='none', tracking=True)


    ingreso_fecha = fields.Datetime(
        string='Fecha de ingreso',
        tracking=True,
        help='Fecha y hora en que se confirmó el ingreso mediante el scanner.'
    )

    ingreso_fuente = fields.Selection(
        [
            ('qr', 'QR / Código de barras'),
            ('ocr', 'OCR (foto)'),
            ('manual', 'Manual'),
        ],
        string='Fuente de ingreso',
        tracking=True,
        help='Origen del registro de ingreso (QR, OCR o manual).'
    )
    @api.onchange('descripcion')
    def _onchange_descripcion(self):
        _logger.debug('Onchange Description for Record IDs: %s', self.ids)
        # Verifica cada registro en el conjunto
        for record in self:
            if record.descripcion:
                record.activador = 'si'
            else:
                record.activador = 'no'
        # Registra el estado final después de evaluar la descripción
        _logger.debug('Activador set to: %s', ', '.join([rec.activador for rec in self]))

    contometro = fields.Char(string='Contometro', required=True, tracking=True,)
    ultima_actualizacion_snmp = fields.Datetime(
        string='Última Actualización SNMP',
        readonly=True,
        help='Fecha y hora de la última vez que SNMP actualizó el contador',
        copy=False
    )

    contador_antes_snmp = fields.Char(
        string='Contador Previo a SNMP',
        readonly=True,
        help='Valor del contador antes de la última actualización SNMP',
        copy=False
    )

    total_actualizaciones_snmp = fields.Integer(
        string='Total Actualizaciones SNMP',
        readonly=True,
        default=0,
        help='Cantidad de veces que SNMP ha actualizado esta máquina',
        copy=False
    )

    ultima_fuente_actualizacion = fields.Selection([
        ('manual', 'Manual'),
        ('snmp', 'SNMP'),
        ('reparacion', 'Reparación'),
    ], string='Última Fuente', readonly=True, default='manual', copy=False)
    # ==========================================================
    # CONTADORES SNMP DETALLADOS - ÚLTIMA LECTURA
    # ==========================================================

    snmp_contador_total = fields.Integer(
        string='SNMP Total',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_bn = fields.Integer(
        string='SNMP B/N',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_color = fields.Integer(
        string='SNMP Color',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_copias = fields.Integer(
        string='SNMP Copias',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_impresiones = fields.Integer(
        string='SNMP Impresiones',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_scanner = fields.Integer(
        string='SNMP Scanner',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_duplex = fields.Integer(
        string='SNMP Dúplex',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_copias_bn = fields.Integer(
        string='SNMP Copias B/N',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_impresiones_bn = fields.Integer(
        string='SNMP Impresiones B/N',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_copias_color = fields.Integer(
        string='SNMP Copias Color',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_impresiones_color = fields.Integer(
        string='SNMP Impresiones Color',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_fax = fields.Integer(
        string='SNMP Fax',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_contador_gran_total = fields.Integer(
        string='SNMP Gran Total',
        readonly=True,
        copy=False,
        tracking=True,
    )

    # ==========================================================
    # TÓNER SNMP - ÚLTIMA LECTURA
    # ==========================================================

    snmp_toner_negro = fields.Float(
        string='SNMP Tóner Negro (%)',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_toner_cyan = fields.Float(
        string='SNMP Tóner Cyan (%)',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_toner_magenta = fields.Float(
        string='SNMP Tóner Magenta (%)',
        readonly=True,
        copy=False,
        tracking=True,
    )

    snmp_toner_amarillo = fields.Float(
        string='SNMP Tóner Amarillo (%)',
        readonly=True,
        copy=False,
        tracking=True,
    )
    def _sat_prepare_snmp_detail_vals_from_context(self):
        """
        Prepara valores SNMP detallados para guardar en sat.sat
        usando el payload/counters/toner que llega por context.
        """
        self.ensure_one()

        snmp_payload = self.env.context.get('snmp_payload') or {}
        snmp_counters = self.env.context.get('snmp_counters') or {}
        snmp_toner = self.env.context.get('snmp_toner') or {}

        if not isinstance(snmp_payload, dict):
            snmp_payload = {}

        if not isinstance(snmp_counters, dict):
            snmp_counters = {}

        if not isinstance(snmp_toner, dict):
            snmp_toner = {}

        counters = snmp_payload.get('counters') if isinstance(snmp_payload.get('counters'), dict) else snmp_counters
        toner = snmp_payload.get('toner') if isinstance(snmp_payload.get('toner'), dict) else snmp_toner

        counters = counters or {}
        toner = toner or {}

        def _num(*keys):
            for key in keys:
                value = None

                if key in snmp_payload and snmp_payload.get(key) not in (None, False, ''):
                    value = snmp_payload.get(key)
                elif key in counters and counters.get(key) not in (None, False, ''):
                    value = counters.get(key)

                if value not in (None, False, ''):
                    try:
                        text = str(value).replace(',', '').strip()
                        match = re.search(r'-?\d+', text)
                        return int(match.group(0)) if match else 0
                    except Exception:
                        return 0

            return 0

        def _float_toner(*keys):
            for key in keys:
                value = None

                if key in toner and toner.get(key) not in (None, False, ''):
                    value = toner.get(key)
                elif key in snmp_payload and snmp_payload.get(key) not in (None, False, ''):
                    value = snmp_payload.get(key)

                if value not in (None, False, ''):
                    try:
                        text = str(value).replace(',', '').strip()
                        match = re.search(r'-?\d+(?:\.\d+)?', text)
                        return float(match.group(0)) if match else 0.0
                    except Exception:
                        return 0.0

            return 0.0

        return {
            'snmp_contador_total': _num('total_counter', 'grand_total_counter', 'total', 'total_counter'),
            'snmp_contador_bn': _num('bw_counter', 'bw', 'bw_total', 'bn'),
            'snmp_contador_color': _num('color_counter', 'color', 'color_total', 'full_color'),

            'snmp_contador_copias': _num('copy_counter', 'copy', 'copy_total', 'copies'),
            'snmp_contador_impresiones': _num('print_counter', 'print', 'print_total', 'prints'),
            'snmp_contador_scanner': _num('scan_counter', 'scan', 'scanner', 'scans'),
            'snmp_contador_duplex': _num('duplex_counter', 'duplex', 'duplex_total'),

            'snmp_contador_copias_bn': _num('copy_bw', 'copies_bw', 'copy_black'),
            'snmp_contador_impresiones_bn': _num('print_bw', 'prints_bw', 'print_black'),

            'snmp_contador_copias_color': _num('copy_color', 'copies_color'),
            'snmp_contador_impresiones_color': _num('print_color', 'prints_color', 'print_full_color'),

            'snmp_contador_fax': _num('fax_counter', 'fax'),
            'snmp_contador_gran_total': _num('grand_total_counter', 'gran_total', 'total_counter', 'total'),

            'snmp_toner_negro': _float_toner('black', 'k', 'negro'),
            'snmp_toner_cyan': _float_toner('cyan', 'c'),
            'snmp_toner_magenta': _float_toner('magenta', 'm'),
            'snmp_toner_amarillo': _float_toner('yellow', 'amarillo', 'y'),
        }
    contometro_proveedor = fields.Char(
        string='Contómetro proveedor (llegada)',
        tracking=True,
        help='Valor declarado por el proveedor/hoja al recibir el equipo. No debe ser sobrescrito por SNMP.'
    )

    alerta_proveedor_snmp_enviada = fields.Boolean(
        string='Alerta proveedor vs SNMP enviada',
        default=False,
        copy=False
    )
    last_snmp_counter_whatsapp = fields.Char(
        string='Último contómetro notificado por WhatsApp (SNMP)',
        readonly=True,
        copy=False
    )

    last_snmp_whatsapp_at = fields.Datetime(
        string='Última alerta WhatsApp SNMP',
        readonly=True,
        copy=False
    )
    


    def _get_reparacion_activa_para_alerta_snmp(self):
        """Retorna la reparación activa (en_revision) más reciente para esta máquina."""
        self.ensure_one()
        Reparacion = self.env['reparaciones.reparaciones']
        rep = Reparacion.search([
            ('maquina_id', '=', self.id),
            ('estado_id', '=', 'en_revision'),
        ], order='create_date desc, id desc', limit=1)
        return rep or False


    

    marca = fields.Char(string='Marca', related='name.marca_id.name', readonly=True, store=True, tracking=True
                        )
    precio_venta = fields.Float(string='Precio de venta', related='name.precio_venta', readonly=True, tracking=True)
    tipo_id = fields.Selection([('color', 'Color'), ('monocromatica', 'Monocromatica')],
                               string='Tipo de maquina', related='name.tipo_id', readonly=True, tracking=True)
    
    tipo_maquina = fields.Char(related='name.tipo_maquina_id.name', readonly=True, string='Tipo de maquina', tracking=True)

    asesora_mobile_clean = fields.Char(
    string='Número de celular asesora (limpio)',
    compute='_compute_asesora_mobile_clean',
    store=True
)
    foto_problema = fields.Binary(
        string='Foto problema',
    )
    
    @api.depends('cliente_id.asesora_id.mobile')
    def _compute_asesora_mobile_clean(self):
        for record in self:
            if record.cliente_id.asesora_id.mobile:
                phone = record.cliente_id.asesora_id.mobile.replace('+', '')
                phone = ''.join(phone.split())
                if not phone.startswith('51'):
                    phone = '51' + phone
                record.asesora_mobile_clean = phone
            else:
                record.asesora_mobile_clean = ''

    



    
    @api.depends('cliente_id', 'estado_ventas_id')
    def _compute_disponibilidad_id(self):
        for record in self:
            old_disponibilidad = record.disponibilidad_id
            _logger.debug('Computing Disponibilidad para ID: %s', record.id)

            if record.estado_ventas_id in ['sin_revisar', 'en_revision', 'finalizado', 'para_revision'] and record.cliente_id:
                _logger.debug('Estado requiere separación, actualizando disponibilidad a separada para ID: %s', record.id)
                record.disponibilidad_id = 'separada'

                if old_disponibilidad != 'separada':
                    record.fecha_separacion = fields.Date.today()
                    _logger.info(f"[TOKEN AUTO] Máquina ID {record.id} separada. Fecha separacion: {record.fecha_separacion}")

                    # 🚨 Generar token si está en ubicación que lo requiere
                    if record.ubicacion_id in ['segundo_local', 'covida'] and not record.location_change_token:
                        _logger.info(f"[TOKEN AUTO] Generando token automáticamente para ID {record.id} en ubicación {record.ubicacion_id}")
                        record.generate_location_change_token()

            elif record.estado_ventas_id in ['sin_revisar', 'en_revision', 'finalizado', 'para_revision'] and not record.cliente_id:
                _logger.debug('Cliente no asignado, disponibilidad = disponible para ID: %s', record.id)
                record.disponibilidad_id = 'disponible'
                record.fecha_separacion = False

            else:
                _logger.debug('Estado fuera de revisión. Disponibilidad = no_disponible para ID: %s', record.id)
                record.disponibilidad_id = 'no_disponible'
                record.fecha_separacion = False

            _logger.debug('Disponibilidad ID final: %s para ID: %s', record.disponibilidad_id, record.id)


    @api.onchange('factura_venta')
    def onchange_factura_venta(self):
        if self.factura_venta:
            self.estado_ventas_id = 'entregada'
            
            
    
    def generate_record_url(self, record):
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        action_id = self.env.ref('sat.action_window').id
        menu_id = self.env.ref('sat.stock_maquinas').id
        url = "{}/web#id={}&view_type=form&model=sat.sat&action={}&menu_id={}".format(base_url, record.id, action_id, menu_id)
        return url
    qr_image = fields.Binary("QR Image", compute="generate_qr_code", attachment=True, store=True)


    @api.depends('estado_ventas_id')  # Suponiendo que quieras codificar un campo específico, reemplaza 'nombre_del_campo_a_codificar' con el campo relevante.
    def generate_qr_code(self):
        for record in self:
            # Generar la URL del registro
            url = self.generate_record_url(record)
            
            if not url:
                continue
            
            # Crear un objeto de código QR
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(url)
            qr.make(fit=True)
            
            # Generar la imagen del código QR
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Guardar la imagen en un buffer en memoria
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            
            # Codificar la imagen en base64
            img_base64 = base64.b64encode(buffer.read())
            
            # Guardar la imagen codificada en el campo qr_image
            record.qr_image = img_base64
    icono_rojo = fields.Html(compute='_compute_icono_rojo', string=' ')

    @api.depends('activador')  # Reemplaza con el campo real que afecta la condición
    def _compute_icono_rojo(self):
        for record in self:
            # Aquí va la lógica para determinar cuándo el ícono debe ser rojo
            if record.activador:  # Reemplaza con tu condición real
                record.icono_rojo = '<i class="fa fa-exclamation-circle icono-grande" style="color: red;"></i>'
            else:
                record.icono_rojo = False

    def action_primer(self):
        self.ubicacion_id = 'primer_piso'

    def action_tercero(self):
        self.ubicacion_id = 'tercer_piso'

    def action_segundo(self):
        self.ubicacion_id = 'segundo_local'

    def create_reparacion(self):
        self.ensure_one()  # Asegurarse de que se trabaja con un único registro.
        return {
            'name': 'Crear Reparación',
            'type': 'ir.actions.act_window',
            'res_model': 'reparaciones.reparaciones',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_maquina_id': self.id,
                
            },
        }

    fecha_para_revision = fields.Datetime(string="Fecha para Revisión", readonly=True, traking=True, store=True)

    posicion_cola = fields.Integer(
        string="Puesto en cola",
        compute='_compute_posicion_cola',
        store=False
    )
    def _compute_posicion_cola(self):
        """
        Calcula el puesto en la cola de revisión según fecha_para_revision.
        Solo considera máquinas en estado 'para_revision' o 'en_revision'
        con fecha_para_revision definida.
        """
        for record in self:
            # Si no está en revisión o no tiene fecha, no tiene puesto
            if record.estado_ventas_id not in ['para_revision', 'en_revision'] or not record.fecha_para_revision:
                record.posicion_cola = 0
                continue

            domain = [
                ('estado_ventas_id', 'in', ['para_revision', 'en_revision']),
                ('fecha_para_revision', '!=', False),
            ]
            # Ordenamos por fecha de revisión e id para tener un orden estable
            cola = self.search(domain, order='fecha_para_revision asc, id asc')
            posicion = 0
            for idx, rec in enumerate(cola, start=1):
                if rec.id == record.id:
                    posicion = idx
                    break

            record.posicion_cola = posicion

    def _sat_safe_postprocess(self, label, callback):
        """
        Ejecuta post-procesos no críticos sin dejar abortada la transacción.
        Si una notificación/cálculo auxiliar falla en PostgreSQL, el savepoint
        revierte solo ese bloque y permite que el write principal continúe.
        """
        self.ensure_one()
        try:
            with self.env.cr.savepoint():
                return callback()
        except Exception as e:
            _logger.error(
                "[SAT POST] Error en %s | ID=%s | error=%s",
                label,
                self.id,
                e,
                exc_info=True,
            )
            return False

    partes_retiradas_ids = fields.One2many(
        'solicitud.parte.tecnico.linea', 'maquina_origen_sat_id',
        string='Partes Retiradas', readonly=True)
    def action_colocar_en_revision(self):
        """
        Coloca la máquina en estado 'para_revision',
        registra fecha_para_revision (en UTC, naïve para Odoo)
        y muestra el puesto en cola.
        """
        self.ensure_one()

        # Ya NO obligamos a tener tipo_revision/prioridad
        # Son completamente opcionales ahora.

        # Fecha/hora Lima usando tu helper
        peru_dt = self._get_peru_datetime()
        # Convertimos a UTC
        utc_dt = peru_dt.astimezone(pytz.utc)
        # Odoo quiere datetime naive o string sin tz -> usamos string como en tu write anterior
        fecha_str = utc_dt.strftime('%Y-%m-%d %H:%M:%S')

        # Cambiar estado y fecha
        self.write({
            'estado_ventas_id': 'para_revision',
            'fecha_para_revision': fecha_str,
        })

        # Recalcular puesto en cola
        self._compute_posicion_cola()
        puesto = self.posicion_cola or 1

        # Notificación en chatter - usar formato HTML correcto
        isidro_partner_id = self.get_isidro_partner_id()
        mensaje_chatter = (
            "Se ha colocado una nueva máquina para revisión.\n\n"
            "Detalles del equipo:\n"
            f"- Modelo: {self.name.name if self.name else ''}\n"
            f"- Serie: {self.serie_id or ''}\n"
            f"- Fecha de registro: {peru_dt.strftime('%Y-%m-%d %H:%M:%S')} (hora Lima)\n"
            f"- Puesto en cola: {puesto}\n\n"
            f"Modificado por: {self.env.user.name}"
        )

        
        self._sat_safe_postprocess(
            "chatter colocar en revision",
            lambda: self.message_post(
                body=mensaje_chatter,
                partner_ids=[isidro_partner_id] if isidro_partner_id else None,
                subtype_xmlid='mail.mt_comment',
            ),
        )

        # Notificación visual al usuario
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Máquina en revisión"),
                'message': _(
                    "La máquina ha sido colocada en la cola de revisión.\n"
                    "Puesto actual: %(puesto)s"
                ) % {'puesto': puesto},
                'type': 'success',
                'sticky': False,
            }
        }

    def action_quitar_de_revision(self):
        """
        Quita la(s) máquina(s) de la cola de revisión,
        limpia fecha_para_revision y deja el estado en 'sin_revisar'
        (ajusta si quieres otro estado final).
        """
        self.ensure_one()

        vals = {
            'fecha_para_revision': False,
        }

        # Solo volver a 'sin_revisar' si actualmente estaba para revisión
        if self.estado_ventas_id == 'para_revision':
            vals['estado_ventas_id'] = 'sin_revisar'

        self.write(vals)

        # Mensaje en chatter
        self.message_post(
            body=_("La máquina fue retirada de la cola de revisión."),
            subtype_xmlid='mail.mt_note',
        )

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Revisión cancelada"),
                'message': _(
                    "La máquina ha sido retirada de la cola de revisión."
                ),
                'type': 'warning',
                'sticky': False,
            }
        }

    def get_isidro_partner_id(self):
        isidro_user = self.env['res.users'].search([('name', '=', 'Isidro Vera Polo')], limit=1)
        if isidro_user:
            return isidro_user.partner_id.id
        return False

    def _get_peru_datetime(self):
        """Obtiene la fecha y hora actual en la zona horaria de Lima/Perú"""
        peru_tz = pytz.timezone('America/Lima')
        utc_now = pytz.utc.localize(datetime.utcnow())
        return utc_now.astimezone(peru_tz)

    def write(self, vals):
        """
        WRITE PROTEGIDO / REFACTORIZADO:
        - Mantiene la lógica actual.
        - Separa la lógica pesada en métodos privados del mismo modelo.
        - Corrige el caso:
            con_problemas -> de_partes
          para que NO mande correo de disponibilidad.
        - No requiere modificar otros archivos.
        """

        # ---------------------------
        # 0) ANTI-RECURSIÓN / ANTI-SPAM
        # ---------------------------
        if self._sat_is_internal_write(vals):
            return super(SatSat, self).write(vals)

        # ---------------------------
        # 1) Preparar vals
        # ---------------------------
        vals = dict(vals or {})
        vals_ingreso_in = self._sat_get_ingreso_vals(vals)

        _logger.error(
            "SAT.WRITE → ENTRANDO | IDS=%s | VALS=%s | INGRESO_IN=%s",
            self.ids,
            vals,
            vals_ingreso_in,
        )

        # Si corrigen el contador del proveedor, permitir alertar de nuevo
        if 'contometro_proveedor' in vals:
            vals.setdefault('alerta_proveedor_snmp_enviada', False)
            vals.setdefault('last_snmp_counter_whatsapp', False)
            vals.setdefault('last_snmp_whatsapp_at', False)

        estados_problema = ['con_problemas', 'de_partes']
        estado_final_no_notificar = 'entregada'

        # ---------------------------
        # 2) Snapshot antes del super()
        # ---------------------------
        cambios_previos = self._sat_get_write_snapshot()

        # ---------------------------
        # 3) Plan de notificaciones por cambio de estado
        # ---------------------------
        estado_plan = self._sat_get_estado_notification_plan(
            vals=vals,
            estados_problema=estados_problema,
            estado_final_no_notificar=estado_final_no_notificar,
        )

        problem_notification_ids = estado_plan.get('problem_notification_ids', set())
        availability_notification_ids = estado_plan.get('availability_notification_ids', set())
        clean_description_ids = estado_plan.get('clean_description_ids', set())

        # Si sale de problema/partes hacia estado operativo, limpiar descripción.
        # En formulario normal Odoo escribe un solo registro.
        if clean_description_ids:
            vals['descripcion'] = False
            vals['activador'] = 'no'

        # ---------------------------
        # 4) Protección final ingreso scanner
        # ---------------------------
        if vals_ingreso_in:
            vals.update(vals_ingreso_in)
        # Si viene de SNMP, guardar también contadores detallados en sat.sat
        if vals.get('ultima_fuente_actualizacion') == 'snmp':
            for record in self:
                try:
                    with record.env.cr.savepoint():
                        snmp_detail_vals = record._sat_prepare_snmp_detail_vals_from_context()
                        vals.update(snmp_detail_vals)

                        _logger.info(
                            "[SAT SNMP DETALLE] Valores detallados preparados para máquina ID=%s | %s",
                            record.id,
                            snmp_detail_vals,
                        )
                except Exception as e:
                    _logger.error(
                        "[SAT SNMP DETALLE] Error preparando contadores detallados SNMP | ID=%s | error=%s",
                        record.id,
                        e,
                        exc_info=True,
                    )
        _logger.error(
            "SAT.WRITE → ANTES SUPER | IDS=%s | VALS=%s | INGRESO_REAPPLY=%s",
            self.ids,
            vals,
            vals_ingreso_in,
        )

        # ---------------------------
        # 5) Escritura real
        # ---------------------------
        result = super(SatSat, self).write(vals)

        try:
            self.invalidate_cache()
        except Exception:
            pass

        _logger.error(
            "SAT.WRITE → DESPUÉS SUPER | IDS=%s | check=%s estado=%s fecha=%s fuente=%s",
            self.ids,
            self.mapped("check_ingreso"),
            self.mapped("ingreso_estado"),
            self.mapped("ingreso_fecha"),
            self.mapped("ingreso_fuente"),
        )

        # ---------------------------
        # 6) Post-procesos por registro
        # ---------------------------
        for record in self:
            record._sat_safe_postprocess(
                "transportistas",
                lambda record=record: record._sat_post_write_transportistas(vals),
            )

            record._sat_safe_postprocess(
                "notificaciones estado",
                lambda record=record: record._sat_post_write_estado_notifications(
                    vals=vals,
                    cambios_previos=cambios_previos,
                    problem_notification_ids=problem_notification_ids,
                    availability_notification_ids=availability_notification_ids,
                    clean_description_ids=clean_description_ids,
                ),
            )

            record._sat_safe_postprocess(
                "revision modelo/contador",
                lambda record=record: record._sat_post_write_model_counter_checks(
                    cambios_previos=cambios_previos,
                ),
            )

            record._sat_safe_postprocess(
                "notificaciones snmp",
                lambda record=record: record._sat_post_write_snmp_notifications(
                    cambios_previos=cambios_previos,
                ),
            )

            record._sat_safe_postprocess(
                "check ingreso",
                lambda record=record: record._sat_post_write_ingreso_check(
                    vals_ingreso_in=vals_ingreso_in,
                ),
            )

            record._sat_safe_postprocess(
                "actualizar prueba snmp",
                lambda record=record: record._sat_post_write_update_prueba_from_snmp(),
            )

        return result

    # ==========================================================
    # HELPERS WRITE
    # ==========================================================

    def _sat_is_internal_write(self, vals):
        """
        Evita reentrada cuando el propio write interno solo actualiza campos técnicos.
        """
        INTERNAL_ONLY_FIELDS = {
            'alerta_proveedor_snmp_enviada',
            'last_snmp_counter_whatsapp',
            'last_snmp_whatsapp_at',
            'location_change_token',
        }

        vals = vals or {}
        return bool(vals) and set(vals.keys()).issubset(INTERNAL_ONLY_FIELDS)

    def _sat_get_ingreso_vals(self, vals):
        """
        Guarda los campos de ingreso scanner que llegaron al write,
        para reinyectarlos antes del super().
        """
        INGRESO_FIELDS = {
            "check_ingreso",
            "ingreso_estado",
            "ingreso_fecha",
            "ingreso_fuente",
        }

        vals = vals or {}
        return {k: vals.get(k) for k in INGRESO_FIELDS if k in vals}

    def _sat_get_write_snapshot(self):
        """
        Snapshot de valores ANTES del super().
        Sirve para comparar cambios después del write.
        """
        cambios_previos = {}

        for record in self:
            cambios_previos[record.id] = {
                'estado_anterior': record.estado_ventas_id,
                'modelo_anterior': record.name.name if record.name else '',
                'tipo_anterior': record.tipo_id,
                'contometro_anterior': record.contometro or '0',
                'fuente_anterior': record.ultima_fuente_actualizacion or '',

                # ingreso antes
                'check_ingreso': bool(record.check_ingreso),
                'ingreso_estado': record.ingreso_estado,
                'ingreso_fecha': record.ingreso_fecha,
                'ingreso_fuente': record.ingreso_fuente,
            }

        return cambios_previos

    def _sat_get_estado_notification_plan(self, vals, estados_problema, estado_final_no_notificar):
        """
        Define qué correos/notificaciones se deben enviar por cambio de estado.

        Reglas:
        - Estado normal -> con_problemas/de_partes:
            enviar problema.
        - con_problemas -> de_partes:
            enviar problema, NO disponibilidad.
        - de_partes -> con_problemas:
            enviar problema, NO disponibilidad.
        - con_problemas/de_partes -> estado normal:
            limpiar descripción y enviar disponibilidad.
        - con_problemas/de_partes -> entregada:
            limpiar descripción, NO enviar disponibilidad.
        """
        problem_notification_ids = set()
        availability_notification_ids = set()
        clean_description_ids = set()

        if 'estado_ventas_id' not in vals:
            return {
                'problem_notification_ids': problem_notification_ids,
                'availability_notification_ids': availability_notification_ids,
                'clean_description_ids': clean_description_ids,
            }

        nuevo_estado = vals.get('estado_ventas_id')

        for record in self:
            estado_anterior = record.estado_ventas_id

            _logger.info(
                "[SAT ESTADO] Evaluando cambio | ID=%s | %s → %s",
                record.id,
                estado_anterior,
                nuevo_estado,
            )

            # A) Entra a problema/partes desde estado normal
            if nuevo_estado in estados_problema and estado_anterior not in estados_problema:
                problem_notification_ids.add(record.id)
                _logger.info(
                    "[SAT ESTADO] Notificar PROBLEMA | ID=%s | %s → %s",
                    record.id,
                    estado_anterior,
                    nuevo_estado,
                )

            # B) Cambio interno problema/partes
            elif estado_anterior in estados_problema and nuevo_estado in estados_problema:
                problem_notification_ids.add(record.id)
                _logger.info(
                    "[SAT ESTADO] Cambio interno problema/partes. "
                    "Notificar PROBLEMA, NO disponibilidad | ID=%s | %s → %s",
                    record.id,
                    estado_anterior,
                    nuevo_estado,
                )

            # C) Sale de problema/partes hacia estado normal
            elif estado_anterior in estados_problema and nuevo_estado not in estados_problema:
                clean_description_ids.add(record.id)

                if nuevo_estado != estado_final_no_notificar:
                    availability_notification_ids.add(record.id)
                    _logger.info(
                        "[SAT ESTADO] Notificar DISPONIBILIDAD | ID=%s | %s → %s",
                        record.id,
                        estado_anterior,
                        nuevo_estado,
                    )
                else:
                    _logger.info(
                        "[SAT ESTADO] Sale de problema hacia ENTREGADA. "
                        "No se notifica disponibilidad | ID=%s | %s → %s",
                        record.id,
                        estado_anterior,
                        nuevo_estado,
                    )

        return {
            'problem_notification_ids': problem_notification_ids,
            'availability_notification_ids': availability_notification_ids,
            'clean_description_ids': clean_description_ids,
        }

    def _sat_post_write_transportistas(self, vals):
        """
        Lógica existente de transporte después del write().
        """
        self.ensure_one()

        try:
            campos_relevantes = {'ubicacion_id', 'cliente_id', 'estado_ventas_id'}

            if campos_relevantes.intersection(vals.keys()):
                if self.disponibilidad_id == 'separada' and self.ubicacion_id in ['segundo_local', 'covida']:
                    with self.env.cr.savepoint():
                        self.enviar_mensaje_transportistas()
                    _logger.info("[TRANSPORTE] Notificación enviada para ID %s", self.id)
                else:
                    _logger.info(
                        "[TRANSPORTE] Sin notificación para ID %s "
                        "(disp=%s, ubic=%s)",
                        self.id,
                        self.disponibilidad_id,
                        self.ubicacion_id,
                    )
            else:
                _logger.info(
                    "[TRANSPORTE] Sin campos relevantes en vals, omitiendo notificación para ID %s",
                    self.id,
                )

        except Exception as e:
            _logger.error("[TRANSPORTE] Error enviando mensaje: %s", e)

    def _sat_post_write_estado_notifications(
        self,
        vals,
        cambios_previos,
        problem_notification_ids,
        availability_notification_ids,
        clean_description_ids,
    ):
        """
        Envía las notificaciones de estado después del write().
        """
        self.ensure_one()

        if self.id in problem_notification_ids:
            try:
                with self.env.cr.savepoint():
                    self.enviar_mensaje_problema_asesora()
                _logger.info(
                    "[SAT ESTADO] Notificación de problema enviada | ID=%s | estado=%s",
                    self.id,
                    self.estado_ventas_id,
                )
            except Exception as e:
                _logger.error(
                    "[SAT ESTADO] Error enviando notificación de problema | ID=%s | error=%s",
                    self.id,
                    e,
                )

        if self.id in availability_notification_ids:
            try:
                with self.env.cr.savepoint():
                    self.enviar_notificacion_disponibilidad()
                _logger.info(
                    "[SAT ESTADO] Notificación de disponibilidad enviada | ID=%s | estado=%s",
                    self.id,
                    self.estado_ventas_id,
                )
            except Exception as e:
                _logger.error(
                    "[SAT ESTADO] Error enviando notificación de disponibilidad | ID=%s | error=%s",
                    self.id,
                    e,
                )

        if self.id in clean_description_ids:
            prev = cambios_previos.get(self.id) or {}
            estado_anterior = prev.get('estado_anterior') or ''
            nuevo_estado = self.estado_ventas_id

            try:
                message = _(
                    "Se limpió la descripción al cambiar el estado de '%s' a '%s'"
                ) % (estado_anterior, nuevo_estado)

                with self.env.cr.savepoint():
                    self.message_post(body=message)

            except Exception as e:
                _logger.error(
                    "[SAT ESTADO] Error publicando mensaje de limpieza | ID=%s | error=%s",
                    self.id,
                    e,
                )

    def _sat_post_write_model_counter_checks(self, cambios_previos):
        """
        Mantiene la revisión de anomalías de modelo y contómetro.
        """
        self.ensure_one()

        prev = cambios_previos.get(self.id) or {}

        modelo_anterior = prev.get('modelo_anterior', '')
        tipo_anterior = prev.get('tipo_anterior')
        contometro_anterior = prev.get('contometro_anterior', '0')

        modelo_nuevo = self.name.name if self.name else ''
        tipo_nuevo = self.tipo_id
        contometro_nuevo = self.contometro or '0'

        # 1) Cambios raros de modelo
        self._check_model_anomalies(
            self,
            modelo_anterior,
            modelo_nuevo,
            tipo_anterior,
            tipo_nuevo,
        )

        # 2) Saltos raros de contómetro
        self._check_counter_anomalies(
            self,
            contometro_anterior,
            contometro_nuevo,
        )

    def _sat_post_write_snmp_notifications(self, cambios_previos):
        """
        Mantiene toda la lógica SNMP existente después del write().
        """
        self.ensure_one()

        prev = cambios_previos.get(self.id) or {}

        modelo_anterior = prev.get('modelo_anterior', '')
        contometro_anterior = prev.get('contometro_anterior', '0')

        modelo_nuevo = self.name.name if self.name else ''
        contometro_nuevo = self.contometro or '0'
        fuente_actual = self.ultima_fuente_actualizacion or ''

        if fuente_actual != 'snmp':
            return

        # 1) Si cambió el modelo
        if modelo_anterior and modelo_nuevo and modelo_anterior != modelo_nuevo:
            try:
                with self.env.cr.savepoint():
                    self.notify_snmp_model_change(
                        previous_model=modelo_anterior,
                        new_model=modelo_nuevo,
                    )
            except Exception as e:
                _logger.error("Error notificando cambio de modelo SNMP: %s", e)

        # 2) Si cambió el contómetro
        if contometro_anterior == contometro_nuevo:
            return

        try:
            old_digits = re.sub(r'[^\d]', '', contometro_anterior or '0')
            new_digits = re.sub(r'[^\d]', '', contometro_nuevo or '0')

            old_val = int(old_digits) if old_digits else 0
            new_val = int(new_digits) if new_digits else 0

            _logger.error(
                "[SNMP DEBUG] ID=%s | old_val=%s new_val=%s | fuente=%s",
                self.id,
                old_val,
                new_val,
                fuente_actual,
            )

            # =====================================================
            # CASO 1: ALERTA Proveedor vs SNMP
            # WhatsApp + Correo, solo una vez
            # =====================================================
            prov_val = 0
            try:
                prov_val = self._to_int_digits(self.contometro_proveedor)
            except Exception as e:
                _logger.error("[SNMP DEBUG] Error obteniendo prov_val: %s", e)
                prov_val = 0

            ya_alertado = bool(self.alerta_proveedor_snmp_enviada)
            prov_alert_sent_this_cycle = False

            _logger.error(
                "[SNMP DEBUG] prov_val=%s | ya_alertado=%s | contometro_proveedor='%s'",
                prov_val,
                ya_alertado,
                self.contometro_proveedor,
            )

            if prov_val and self._is_proveedor_vs_snmp_alert(prov_val, new_val):
                if not ya_alertado:
                    _logger.error(
                        "[SNMP ALERT] ✅ ENVIANDO alerta Proveedor vs SNMP. prov=%s snmp=%s | ID=%s",
                        prov_val,
                        new_val,
                        self.id,
                    )

                    # WhatsApp técnico
                    try:
                        with self.env.cr.savepoint():
                            self._notify_tecnico_guardar_hoja_contometro_snmp(prov_val, new_val)
                    except Exception as e:
                        _logger.error("[SNMP->WA] Error alertando al técnico: %s", e)

                    # Correo anomalía
                    try:
                        with self.env.cr.savepoint():
                            self.notify_snmp_counter_update(
                                previous_counter=prov_val,
                                new_counter=new_val,
                                is_anomaly=True,
                            )
                    except Exception as e:
                        _logger.error("[SNMP->MAIL] Error enviando correo proveedor vs SNMP: %s", e)

                    prov_alert_sent_this_cycle = True

                    # Marcar como alertado
                    try:
                        with self.env.cr.savepoint():
                            self.sudo().write({'alerta_proveedor_snmp_enviada': True})
                    except Exception as e:
                        _logger.error("[SNMP ALERT] Error marcando alerta_proveedor_snmp_enviada: %s", e)

                else:
                    _logger.error(
                        "[SNMP ALERT] ❌ Ya alertado antes. NO repetir | ID=%s",
                        self.id,
                    )
            else:
                _logger.error(
                    "[SNMP] ℹ️ Proveedor vs SNMP normal o sin proveedor. prov=%s snmp=%s | ID=%s",
                    prov_val,
                    new_val,
                    self.id,
                )

            # =====================================================
            # CASO 2: Anomalía técnica old -> new
            # Solo si no se envió alerta proveedor en este ciclo
            # =====================================================
            _logger.error(
                "[SNMP DEBUG] Evaluando CASO 2 | prov_alert_sent=%s",
                prov_alert_sent_this_cycle,
            )

            if not prov_alert_sent_this_cycle:
                is_anomaly = self._is_counter_anomaly(old_val, new_val)

                _logger.error(
                    "[SNMP DEBUG] _is_counter_anomaly(%s, %s) = %s",
                    old_val,
                    new_val,
                    is_anomaly,
                )

                if is_anomaly:
                    _logger.error(
                        "[SNMP ANOMALY] ✅ ENVIANDO correo técnico old->new: %s → %s | ID=%s",
                        old_val,
                        new_val,
                        self.id,
                    )

                    try:
                        with self.env.cr.savepoint():
                            self.notify_snmp_counter_update(
                                previous_counter=old_val,
                                new_counter=new_val,
                                is_anomaly=True,
                            )
                        _logger.error("[SNMP ANOMALY] ✅ Correo enviado exitosamente")
                    except Exception as e:
                        _logger.error("[SNMP ANOMALY] ❌ Error enviando correo: %s", e)
                else:
                    _logger.error(
                        "[SNMP] ℹ️ Cambio normal %s → %s, sin correo | ID=%s",
                        old_val,
                        new_val,
                        self.id,
                    )
            else:
                _logger.error(
                    "[SNMP] ℹ️ Omitiendo CASO 2 porque ya se envió alerta Proveedor | ID=%s",
                    self.id,
                )

        except Exception as e:
            _logger.error("[SNMP ERROR] Error procesando contador: %s", e, exc_info=True)

    def _sat_post_write_ingreso_check(self, vals_ingreso_in):
        """
        Log extra para verificar que los campos de ingreso scanner no se pierdan.
        """
        self.ensure_one()

        if not vals_ingreso_in:
            return

        try:
            self.invalidate_cache()
        except Exception:
            pass

        _logger.error(
            "SAT.WRITE → CHECK INGRESO POST | ID=%s | IN=%s | NOW=(check=%s estado=%s fecha=%s fuente=%s)",
            self.id,
            vals_ingreso_in,
            bool(self.check_ingreso),
            self.ingreso_estado,
            self.ingreso_fecha,
            self.ingreso_fuente,
        )

    def _sat_post_write_update_prueba_from_snmp(self):
        """
        Actualiza la prueba técnica desde SNMP.

        Flujo:
        - Solo corre cuando ultima_fuente_actualizacion = 'snmp'.
        - Busca la última prueba técnica de la máquina.
        - Toma del context:
            * snmp_counters
            * snmp_toner
            * snmp_payload
        - Asegura que el payload completo no se pierda.
        - Si el controlador solo manda counters/toner, arma payload básico.
        - Si el controlador manda units/supplies/raw_units/raw_supplies,
        los conserva para que aplicar_snmp_payload() los guarde y luego
        el dashboard los pueda mostrar.
        """
        self.ensure_one()

        try:
            # Solo procesar cuando la actualización viene realmente de SNMP.
            if self.ultima_fuente_actualizacion != 'snmp':
                return

            with self.env.cr.savepoint():
                Prueba = self.env['sat.prueba.maquina'].sudo()

                prueba = Prueba.search([
                    ('maquina_id', '=', self.id),
                ], order='id desc', limit=1)

            if not prueba:
                _logger.warning(
                    "[PRUEBA SNMP] No se encontró prueba activa para máquina ID %s | Serie=%s",
                    self.id,
                    self.serie_id,
                )
                return

            snmp_counters = self.env.context.get('snmp_counters') or {}
            snmp_toner = self.env.context.get('snmp_toner') or {}
            snmp_payload = self.env.context.get('snmp_payload') or {}

            if not isinstance(snmp_counters, dict):
                snmp_counters = {}

            if not isinstance(snmp_toner, dict):
                snmp_toner = {}

            if not isinstance(snmp_payload, dict):
                snmp_payload = {}

            # ======================================================
            # Si no llega payload completo, construir uno básico.
            # ======================================================
            if not snmp_payload:
                snmp_payload = {
                    'ip': False,
                    'serial': self.serie_id,
                    'brand': self.marca or '',
                    'model': self.name.name if self.name else '',
                    'total_counter': self.contometro,
                    'counters': snmp_counters,
                    'toner': snmp_toner,
                }

            # ======================================================
            # Asegurar que counters y toner estén dentro del payload.
            # No usar reemplazo destructivo, solo completar.
            # ======================================================
            if isinstance(snmp_counters, dict) and snmp_counters:
                if not isinstance(snmp_payload.get('counters'), dict):
                    snmp_payload['counters'] = {}
                snmp_payload['counters'].update(snmp_counters)

            if isinstance(snmp_toner, dict) and snmp_toner:
                if not isinstance(snmp_payload.get('toner'), dict):
                    snmp_payload['toner'] = {}
                snmp_payload['toner'].update(snmp_toner)

            # ======================================================
            # Completar datos básicos sin borrar lo que ya vino.
            # ======================================================
            if not snmp_payload.get('total_counter'):
                snmp_payload['total_counter'] = self.contometro

            snmp_payload.setdefault('serial', self.serie_id)
            snmp_payload.setdefault('brand', self.marca or '')
            snmp_payload.setdefault('model', self.name.name if self.name else '')

            # ======================================================
            # Normalización suave de bloques comunes.
            # Esto NO inventa datos, solo replica nombres para que
            # aplicar_snmp_payload/dashboard tengan más probabilidad de leerlos.
            # ======================================================

            # supplies / consumables
            if isinstance(snmp_payload.get('consumables'), dict) and not snmp_payload.get('supplies'):
                snmp_payload['supplies'] = snmp_payload.get('consumables')

            if isinstance(snmp_payload.get('raw_consumables'), dict) and not snmp_payload.get('raw_supplies'):
                snmp_payload['raw_supplies'] = snmp_payload.get('raw_consumables')

            # units / maintenance / life
            if isinstance(snmp_payload.get('maintenance'), dict) and not snmp_payload.get('units'):
                snmp_payload['units'] = snmp_payload.get('maintenance')

            if isinstance(snmp_payload.get('life'), dict) and not snmp_payload.get('units'):
                snmp_payload['units'] = snmp_payload.get('life')

            if isinstance(snmp_payload.get('lifetime'), dict) and not snmp_payload.get('units'):
                snmp_payload['units'] = snmp_payload.get('lifetime')

            # Bloques específicos hacia units si no existe units
            unidades_compuestas = {}

            for key in [
                'developer',
                'developers',
                'drum',
                'drums',
                'fuser',
                'fusers',
                'fixing',
                'transfer',
                'transfer_belt',
                'image_unit',
                'imaging_unit',
            ]:
                value = snmp_payload.get(key)
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        unidades_compuestas['%s_%s' % (key, sub_key)] = sub_value
                elif value not in (None, False, ''):
                    unidades_compuestas[key] = value

            if unidades_compuestas:
                if not isinstance(snmp_payload.get('units'), dict):
                    snmp_payload['units'] = {}
                snmp_payload['units'].update(unidades_compuestas)

            # Waste toner hacia supplies
            consumibles_compuestos = {}

            for key in [
                'waste',
                'waste_toner',
                'wasteToner',
            ]:
                value = snmp_payload.get(key)
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        consumibles_compuestos['%s_%s' % (key, sub_key)] = sub_value
                elif value not in (None, False, ''):
                    consumibles_compuestos[key] = value

            if consumibles_compuestos:
                if not isinstance(snmp_payload.get('supplies'), dict):
                    snmp_payload['supplies'] = {}
                snmp_payload['supplies'].update(consumibles_compuestos)

            # ======================================================
            # Aplicar payload a la prueba.
            # Este método debe guardar raw_payload_json y snmp_detalle_ids.
            # ======================================================
            with self.env.cr.savepoint():
                prueba.aplicar_snmp_payload(
                    counters=snmp_counters,
                    toner=snmp_toner,
                    payload=snmp_payload,
                )

            _logger.info(
                "[PRUEBA SNMP] Payload aplicado en prueba ID %s | Máquina ID=%s | Serie=%s | counters=%s | toner=%s | payload_keys=%s",
                prueba.id,
                self.id,
                self.serie_id,
                list(snmp_counters.keys()) if isinstance(snmp_counters, dict) else [],
                list(snmp_toner.keys()) if isinstance(snmp_toner, dict) else [],
                list(snmp_payload.keys()) if isinstance(snmp_payload, dict) else [],
            )

        except Exception as e:
            _logger.error(
                "[PRUEBA SNMP] Error actualizando prueba desde SNMP para máquina ID %s: %s",
                self.id,
                e,
                exc_info=True,
            )
    
    prueba_ids = fields.One2many(
        'sat.prueba.maquina',
        'maquina_id',
        string='Pruebas técnicas'
    )
    def _is_counter_anomaly(self, old_val, new_val):
        """
        Detecta si un cambio de contador es anómalo para RECLAMAR al proveedor.
        
        Maneja CUALQUIER valor de contador (desde 1 hasta millones).
        
        NOTIFICA cuando (SOLO INCREMENTOS):
        1. SNMP tiene MÁS dígitos (aunque sea +1 dígito) → SIEMPRE RECLAMAR
        2. Mismos dígitos pero incremento ≥ 20,000 → RECLAMAR
        
        NO notifica cuando:
        - Decremento (contador baja) → No tiene sentido reclamar
        - SNMP tiene MENOS dígitos → No tiene sentido reclamar
        - Incremento normal < 20,000 con mismos dígitos → Uso normal
        
        Ejemplos:
        - 40,000 → 35,000 (decremento) → ❌ NO
        - 40,000 → 4,000 (menos dígitos) → ❌ NO
        - 40,000 → 42,000 (+2,000) → ❌ NO
        - 40,000 → 60,000 (+20,000) → ✅ SÍ (RECLAMAR)
        - 40,000 → 400,000 (más dígitos) → ✅ SÍ (RECLAMAR)
        - 99 → 100 (+1 dígito) → ✅ SÍ (RECLAMAR)
        """
        # Si alguno es 0 o negativo, no evaluar
        if old_val <= 0 or new_val <= 0:
            _logger.debug(
                "[SNMP Anomaly] Valores inválidos: old=%s new=%s, NO evaluar",
                old_val, new_val
            )
            return False
        
        # 1) DECREMENTO → NO notificar (no reclamar)
        if new_val < old_val:
            _logger.debug(
                "[SNMP Anomaly] Decremento detectado %s → %s, NO notificar (no reclamar)",
                old_val, new_val
            )
            return False
        
        # 2) Contar dígitos
        old_digits = len(str(int(old_val)))
        new_digits = len(str(int(new_val)))
        
        # 3) MENOS dígitos → NO notificar (no reclamar)
        if new_digits < old_digits:
            _logger.debug(
                "[SNMP Anomaly] Menos dígitos %s → %s (%d → %d dígitos), NO notificar (no reclamar)",
                old_val, new_val, old_digits, new_digits
            )
            return False
        
        # 4) MÁS dígitos (aunque sea +1) → SÍ notificar (RECLAMAR)
        if new_digits > old_digits:
            _logger.info(
                "[SNMP Anomaly] ✅ MÁS dígitos detectado: %s → %s (%d → %d dígitos) - RECLAMAR AL PROVEEDOR",
                old_val, new_val, old_digits, new_digits
            )
            return True
        
        # 5) MISMOS dígitos → Verificar si incremento ≥ 20,000
        diferencia = new_val - old_val
        
        if diferencia >= 20000:
            _logger.info(
                "[SNMP Anomaly] ✅ Mismos dígitos pero incremento ≥ 20,000: %s → %s (+%s) - RECLAMAR AL PROVEEDOR",
                old_val, new_val, diferencia
            )
            return True
        
        # 6) Cambio normal (incremento pequeño)
        _logger.debug(
            "[SNMP] Cambio normal: %s → %s (+%s, %d dígitos), NO notificar",
            old_val, new_val, diferencia, old_digits
        )
        return False


        # =========================================================
    #  NUEVO: helpers para alerta Proveedor vs SNMP (1 sola vez)
    # =========================================================
    def _to_int_digits(self, value):
        """Convierte '12,345' / '12345' / None -> int seguro."""
        try:
            s = re.sub(r"[^\d]", "", str(value or ""))
            return int(s) if s else 0
        except Exception:
            return 0

    def _is_proveedor_vs_snmp_alert(self, prov_val, snmp_val):
        """
        Regla de ALERTA (misma para WhatsApp y Correo) comparando:
        - prov_val: contómetro proveedor (llegada/hoja)
        - snmp_val: contómetro leído por SNMP

        REGLAS:
        1) Solo evaluar si SNMP > proveedor.
        2) Si misma cantidad de dígitos: alertar SOLO si diff >= 20000.
        3) Si SNMP tiene MÁS dígitos que proveedor: alertar SIEMPRE.
        4) Si SNMP tiene MENOS dígitos que proveedor: NO alertar.
        """
        if prov_val <= 0 or snmp_val <= 0:
            return False

        if snmp_val <= prov_val:
            return False

        prov_digits = len(str(int(prov_val)))
        snmp_digits = len(str(int(snmp_val)))

        # Si bajó cantidad de dígitos => NO alertar
        if snmp_digits < prov_digits:
            return False

        # Si subió cantidad de dígitos => ALERTAR siempre
        if snmp_digits > prov_digits:
            return True

        # Misma cantidad de dígitos => alertar si aumenta >= 20000
        diff = snmp_val - prov_val
        return diff >= 20000




    def _check_model_anomalies(self, record, modelo_anterior, modelo_nuevo, tipo_anterior, tipo_nuevo):
        """
        Detecta casos como:
        - Canon 4525  -> 4525/4535 (cambio de velocidad dentro misma familia)
        - bizhub 364e -> bizhub C364e (cambio de mono a color)
        y avisa por chatter/correo (a Isidro).
        """
        # Si no cambió el modelo, no hacemos nada
        if not modelo_anterior or not modelo_nuevo or modelo_anterior == modelo_nuevo:
            return

        # Extraer último bloque numérico de cada modelo (núcleo de velocidad)
        def _get_core_digits(text):
            nums = re.findall(r'\d+', text or '')
            return nums[-1] if nums else None

        core_old = _get_core_digits(modelo_anterior)
        core_new = _get_core_digits(modelo_nuevo)

        posible_cambio_velocidad = False
        detalle_velocidad = ""

        if core_old and core_new and core_old != core_new:
            # Caso típico: 4525 -> 4535 (mismo largo y primeros dígitos iguales)
            try:
                if len(core_old) == len(core_new):
                    if len(core_old) == 4 and core_old[:2] == core_new[:2]:
                        posible_cambio_velocidad = True
                        detalle_velocidad = f"{core_old[-2:]} → {core_new[-2:]}"
                    else:
                        # fallback: cualquier cambio de núcleo con mismo largo
                        posible_cambio_velocidad = True
                        detalle_velocidad = f"{core_old} → {core_new}"
            except Exception:
                pass

        # Cambio de tipo (color/mono)
        cambio_tipo = False
        if tipo_anterior and tipo_nuevo and tipo_anterior != tipo_nuevo:
            cambio_tipo = True

        # Si no hay nada relevante, salir
        if not posible_cambio_velocidad and not cambio_tipo:
            return

        isidro_partner_id = record.get_isidro_partner_id()
        url = record.generate_record_url(record)

        lineas = [
            "Se detectó una actualización relevante del modelo (posiblemente por SNMP o edición manual):",
            f"• Modelo anterior: <b>{modelo_anterior}</b>",
            f"• Modelo nuevo: <b>{modelo_nuevo}</b>",
        ]

        if posible_cambio_velocidad:
            lineas.append(f"• Posible cambio de velocidad (núcleo): <b>{detalle_velocidad}</b>")

        if cambio_tipo:
            # Obtener el selection de tipo_id de forma segura,
            # soportando tanto lista/tupla como función.
            try:
                field_info = record.fields_get(['tipo_id'])['tipo_id']
                sel_tipo = dict(field_info.get('selection', []))
            except Exception:
                sel_tipo = {}

            txt_old = sel_tipo.get(tipo_anterior, tipo_anterior)
            txt_new = sel_tipo.get(tipo_nuevo, tipo_nuevo)
            lineas.append(f"• Cambio de tipo: <b>{txt_old}</b> → <b>{txt_new}</b>")

        lineas.append(f"• Equipo: <b>{record.name.name if record.name else ''}</b> / Serie: <b>{record.serie_id}</b>")
        lineas.append(f"• Enlace al equipo: {url}")

        body = "<br/>".join(lineas)

        record.message_post(
            body=body,
            subtype_xmlid='mail.mt_note',
            partner_ids=[isidro_partner_id] if isidro_partner_id else None,
        )

    def _check_counter_anomalies(self, record, contometro_anterior, contometro_nuevo):
        """
        Detecta variaciones sospechosas en el contómetro PARA RECLAMAR AL PROVEEDOR.
        
        SOLO notifica INCREMENTOS anómalos:
        - Más dígitos (ej: 40,000 → 400,000)
        - Mismos dígitos pero +20,000 o más (ej: 40,000 → 60,000)
        
        NO notifica:
        - Decrementos (no tiene sentido reclamar)
        - Menos dígitos (no tiene sentido reclamar)
        - Incrementos normales < 20,000
        """

        # Limpiar a solo dígitos
        old_digits = re.sub(r'[^\d]', '', contometro_anterior or '') or '0'
        new_digits = re.sub(r'[^\d]', '', contometro_nuevo or '') or '0'

        try:
            old_val = int(old_digits)
            new_val = int(new_digits)
        except Exception:
            return

        # Usar el método de detección de anomalía (ya corregido)
        if not record._is_counter_anomaly(old_val, new_val):
            return

        # Si llegamos aquí, es un incremento anómalo para RECLAMAR
        incremento = new_val - old_val
        isidro_partner_id = record.get_isidro_partner_id()
        url = record.generate_record_url(record)

        old_digit_count = len(str(old_val))
        new_digit_count = len(str(new_val))

        if new_digit_count > old_digit_count:
            # Más dígitos
            lineas = [
                "⚠️ <b>ALERTA: Incremento sospechoso de dígitos en el contómetro</b>",
                f"• Valor anterior: <b>{old_val:,}</b> ({old_digit_count} dígitos)",
                f"• Valor nuevo: <b>{new_val:,}</b> ({new_digit_count} dígitos)",
                f"• Incremento: <b>+{incremento:,}</b>",
                f"• <b>Acción sugerida:</b> Revisar y reclamar al proveedor",
                f"• Equipo: <b>{record.name.name if record.name else ''}</b> / Serie: <b>{record.serie_id}</b>",
                f"• Enlace al equipo: {url}",
            ]
        else:
            # Mismo número de dígitos pero +20,000 o más
            lineas = [
                "⚠️ <b>ALERTA: Incremento inusual en el contómetro</b>",
                f"• Valor anterior: <b>{old_val:,}</b>",
                f"• Valor nuevo: <b>{new_val:,}</b>",
                f"• Incremento: <b>+{incremento:,}</b>",
                f"• <b>Acción sugerida:</b> Revisar y reclamar al proveedor",
                f"• Equipo: <b>{record.name.name if record.name else ''}</b> / Serie: <b>{record.serie_id}</b>",
                f"• Enlace al equipo: {url}",
            ]

        body = "<br/>".join(lineas)

        record.message_post(
            body=body,
            subtype_xmlid='mail.mt_note',
            partner_ids=[isidro_partner_id] if isidro_partner_id else None,
        )


    

    
    # =========================
    #  NOTIFICACIONES SNMP
    # =========================

    def _send_snmp_mail(self, template_xmlid, extra_context=None):
        """
        Helper para enviar correos de SNMP usando plantillas.
        - template_xmlid: ej. 'sat.email_template_snmp_counter_anomaly'
        - extra_context: dict con datos adicionales para usar en la plantilla (ctx)
        """
        self.ensure_one()
        template = self.env.ref(template_xmlid, raise_if_not_found=False)
        if not template:
            _logger.warning("No se encontró la plantilla de correo SNMP: %s", template_xmlid)
            return False

        ctx = dict(self.env.context or {})
        if extra_context:
            ctx.update(extra_context)

        try:
            template.with_context(ctx).sudo().send_mail(self.id, force_send=True)
            _logger.info("Correo SNMP enviado usando plantilla %s para equipo ID %s", template_xmlid, self.id)
            return True
        except Exception as e:
            _logger.error("Error enviando correo SNMP con plantilla %s: %s", template_xmlid, e)
            return False

    def notify_snmp_counter_anomaly(self, old_counter, new_counter, reason=None):
        """
        Notifica cuando hay algo raro en el contómetro:
        - reason = 'decremento'   -> contador bajó
        - reason = 'salto_grande' -> salto muy grande (ej. 2,000 -> 200,000)
        """
        self.ensure_one()

        if reason == 'decremento':
            title = "⚠️ Contador decreció por SNMP"
        else:
            title = "⚠️ Posible inconsistencia en el contador SNMP"

        body = _(
            "%(title)s<br/>Anterior: <b>%(old)s</b><br/>Nuevo: <b>%(new)s</b>"
        ) % {
            'title': title,
            'old': f"{old_counter:,}",
            'new': f"{new_counter:,}",
        }

        # Mensaje en chatter
        self.message_post(body=body)

        # Correo por plantilla (ajusta el XMLID a tus plantillas reales)
        self._send_snmp_mail(
            'sat.email_template_snmp_counter_anomaly',
            {
                'snmp_old_counter': old_counter,
                'snmp_new_counter': new_counter,
                'snmp_reason': reason or '',
            }
        )

    def notify_snmp_model_mismatch(self, snmp_model, current_model, new_model=None):
        """
        Notifica cuando el modelo detectado por SNMP NO coincide con el registrado.
        Ej: MP 3055 vs MP 4055, 4525 vs 4535, etc.
        
        Parámetros:
        - snmp_model: Modelo detectado por SNMP
        - current_model: Modelo que tenía el equipo antes
        - new_model: (Opcional) Modelo al que se cambió automáticamente
        """
        self.ensure_one()

        url = self.generate_record_url(self)
        
        if new_model and new_model != current_model:
            # Caso: Se cambió automáticamente el modelo
            body = _(
                "⚠️ <b>Modelo cambiado automáticamente por SNMP (mismatch de núcleo)</b><br/>"
                "• Modelo anterior: <b>%(cur)s</b><br/>"
                "• Modelo nuevo: <b>%(new)s</b><br/>"
                "• Detectado por SNMP: <b>%(snmp)s</b><br/>"
                "• Serie: <b>%(serie)s</b><br/>"
                "• Enlace: %(url)s"
            ) % {
                'cur': current_model or '—',
                'new': new_model or '—',
                'snmp': snmp_model or '—',
                'serie': self.serie_id or '—',
                'url': url,
            }
        else:
            # Caso: Solo se detectó diferencia, no se cambió
            body = _(
                "⚠️ <b>Diferencia de modelo detectada por SNMP</b><br/>"
                "• Modelo actual: <b>%(cur)s</b><br/>"
                "• Modelo detectado por SNMP: <b>%(snmp)s</b><br/>"
                "• Serie: <b>%(serie)s</b><br/>"
                "• Enlace: %(url)s"
            ) % {
                'cur': current_model or '—',
                'snmp': snmp_model or '—',
                'serie': self.serie_id or '—',
                'url': url,
            }

        self.message_post(body=body)

        self._send_snmp_mail(
            'sat.email_template_snmp_model_mismatch',
            {
                'snmp_current_model': current_model,
                'snmp_detected_model': snmp_model,
                'snmp_new_model': new_model,
                'record_url': url,
            }
        )

    def notify_snmp_model_change(self, previous_model, new_model):
        """
        Notifica cuando SNMP cambió exitosamente el modelo del equipo.
        Esto es para cambios normales, no mismatches.
        """
        self.ensure_one()
        
        url = self.generate_record_url(self)
        
        body = _(
            "✅ <b>Modelo actualizado por SNMP</b><br/>"
            "• Modelo anterior: <b>%(prev)s</b><br/>"
            "• Modelo nuevo: <b>%(new)s</b><br/>"
            "• Serie: <b>%(serie)s</b><br/>"
            "• Enlace: %(url)s"
        ) % {
            'prev': previous_model or '—',
            'new': new_model or '—',
            'serie': self.serie_id or '—',
            'url': url,
        }

        self.message_post(body=body)

        self._send_snmp_mail(
            'sat.email_template_snmp_model_change',
            {
                'snmp_previous_model': previous_model,
                'snmp_new_model': new_model,
                'record_url': url,
            }
        )

    def notify_snmp_counter_update(self, previous_counter, new_counter, is_anomaly=False):
        """
        Notifica cuando SNMP actualizó el contómetro.
        
        Args:
            previous_counter: Valor anterior del contador
            new_counter: Valor nuevo del contador
            is_anomaly: True si es una anomalía (para enviar correo de alerta)
        """
        self.ensure_one()
        
        _logger.error(
            "[CORREO SNMP] INICIANDO notify_snmp_counter_update | ID=%s | prev=%s new=%s | anomaly=%s",
            self.id, previous_counter, new_counter, is_anomaly
        )
        
        # Solo notificar si hay cambio real
        if previous_counter == new_counter:
            _logger.error("[CORREO SNMP] ❌ Valores iguales, abortando")
            return
        
        url = self.generate_record_url(self)
        incremento = new_counter - previous_counter
        
        # Formatear con comas para mejor lectura
        prev_formatted = f"{previous_counter:,}"
        new_formatted = f"{new_counter:,}"
        inc_formatted = f"{abs(incremento):,}"
        
        # Determinar qué plantilla usar
        if is_anomaly:
            # 🚨 ANOMALÍA: Usar plantilla de alerta
            template_xmlid = 'sat.email_template_snmp_counter_anomaly'
            
            # Determinar razón de la anomalía
            if incremento < 0:
                reason = 'decremento'
                emoji = "⬇️"
                tipo = "decreció"
            else:
                reason = 'salto_grande'
                emoji = "⬆️"
                tipo = "aumentó de forma anormal"
            
            body = _(
                "%(emoji)s <b>⚠️ ANOMALÍA: Contómetro %(tipo)s</b><br/>"
                "• Anterior: <b>%(prev)s</b><br/>"
                "• Nuevo: <b>%(new)s</b><br/>"
                "• Cambio: <b>%(inc)s</b><br/>"
                "• Modelo: <b>%(model)s</b><br/>"
                "• Serie: <b>%(serie)s</b><br/>"
                "• Enlace: %(url)s<br/><br/>"
                "<span style='color:red;'>⚠️ Se requiere verificación inmediata del equipo.</span>"
            ) % {
                'emoji': emoji,
                'tipo': tipo,
                'prev': prev_formatted,
                'new': new_formatted,
                'inc': inc_formatted,
                'model': self.name.name if self.name else '—',
                'serie': self.serie_id or '—',
                'url': url,
            }
            
            # Contexto para plantilla de anomalía
            mail_context = {
                'snmp_old_counter': prev_formatted,
                'snmp_new_counter': new_formatted,
                'snmp_reason': reason,
            }
            
        else:
            # ✅ NORMAL: Usar plantilla de actualización normal
            template_xmlid = 'sat.email_template_snmp_counter_update'
            
            if incremento < 0:
                emoji = "⚠️"
                tipo = "decreció"
            else:
                emoji = "✅"
                tipo = "aumentó"
            
            body = _(
                "%(emoji)s <b>Contómetro %(tipo)s por SNMP</b><br/>"
                "• Anterior: <b>%(prev)s</b><br/>"
                "• Nuevo: <b>%(new)s</b><br/>"
                "• Cambio: <b>%(inc)s</b><br/>"
                "• Modelo: <b>%(model)s</b><br/>"
                "• Serie: <b>%(serie)s</b><br/>"
                "• Enlace: %(url)s"
            ) % {
                'emoji': emoji,
                'tipo': tipo,
                'prev': prev_formatted,
                'new': new_formatted,
                'inc': inc_formatted,
                'model': self.name.name if self.name else '—',
                'serie': self.serie_id or '—',
                'url': url,
            }
            
            # Contexto para plantilla normal
            mail_context = {
                'snmp_previous_counter': prev_formatted,
                'snmp_new_counter': new_formatted,
                'snmp_increment': inc_formatted,
                'record_url': url,
            }

        # Mensaje en chatter
        self.message_post(body=body)

        # Enviar correo con la plantilla correcta
        _logger.error("[CORREO SNMP] Llamando _send_snmp_mail con template: %s", template_xmlid)
        
        result = self._send_snmp_mail(template_xmlid, mail_context)
        
        _logger.error("[CORREO SNMP] Resultado _send_snmp_mail: %s", result)
        
        return result

    def notify_snmp_model_suggestion(self, snmp_model):
        """
        Notifica cuando SNMP sugiere un modelo que no existe aún
        o no se pudo asignar automáticamente.
        """
        self.ensure_one()

        url = self.generate_record_url(self)

        body = _(
            "💡 <b>Sugerencia de modelo detectada por SNMP</b><br/>"
            "• Modelo sugerido: <b>%(snmp)s</b><br/>"
            "• Serie: <b>%(serie)s</b><br/>"
            "• Enlace: %(url)s"
        ) % {
            'snmp': snmp_model or '—',
            'serie': self.serie_id or '—',
            'url': url,
        }

        self.message_post(body=body)

        self._send_snmp_mail(
            'sat.email_template_snmp_model_suggestion',
            {
                'snmp_suggested_model': snmp_model,
                'record_url': url,
            }
        )

    
    location_change_token = fields.Char(string="Token de cambio de ubicación", copy=False, readonly=True)
    def generate_location_change_token(self, force=False):
        """Genera un token nuevo si no existe, o si se forza regeneración."""
        import secrets
        for record in self:
            if record.location_change_token and not force:
                _logger.info(f"[TOKEN AUTO] Token ya existe para ID {record.id}, no se regenera: {record.location_change_token}")
                continue

            token = secrets.token_urlsafe(16)
            record.sudo().write({'location_change_token': token})
            token_verificado = self.env['sat.sat'].sudo().browse(record.id).location_change_token

            if token_verificado == token:
                _logger.info(f"[TOKEN AUTO] Token generado y verificado correctamente para ID {record.id}: {token}")
            else:
                _logger.error(f"[TOKEN AUTO] FALLO al guardar token para ID {record.id}. Esperado: {token}, Leído: {token_verificado}")


    
    @api.onchange('disponibilidad_id', 'ubicacion_id')
    def _onchange_disponibilidad_ubicacion(self):
        """
        Onchange solo para UI (NO lógica de negocio).
        Evita enviar mensajes o usar message_post aquí porque el registro
        aún puede ser NewId (no guardado).
        """
        if self.disponibilidad_id == 'separada' and self.ubicacion_id in ['segundo_local', 'covida']:
            return self._notify_vendedora()

    
    
    def crear_url_cambio_ubicacion(self, record):
        import secrets
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        _logger.info(f"[TOKEN] Iniciando generación de token para ID {record.id}")

        # Asegurarse de que no sea un NewId
        if isinstance(record.id, models.NewId):
            # Buscar por serie_id como referencia
            record = self.env['sat.sat'].sudo().search([('serie_id', '=', record.serie_id)], limit=1)
            if not record:
                _logger.error(f"[TOKEN] No se encontró el registro real para NewId: {record.id}")
                return None

        # Continuar con el ID real
        if record.location_change_token:
            token = record.location_change_token
            _logger.info(f"[TOKEN] Ya existe token para ID {record.id}: {token}")
        else:
            token = secrets.token_urlsafe(16)
            record.sudo().write({'location_change_token': token})
            token_verificado = self.env['sat.sat'].sudo().browse(record.id).location_change_token
            if token_verificado != token:
                _logger.error(f"[TOKEN] FALLO al guardar token para ID {record.id}. Esperado: {token}, Leído: {token_verificado}")
                return None
            _logger.info(f"[TOKEN] Token generado correctamente para ID {record.id}: {token}")

        url = f"{base_url}/sat/change_location/{record.id}?token={token}"
        _logger.info(f"[TOKEN] URL final generada: {url}")
        return url



    def invalidar_token_ubicacion(self):
        """Invalida el token de cambio de ubicación después de su uso"""
        try:
            self.write({'location_change_token': False})
            _logger.info(f"Token invalidado para el registro {self.id}")
            return True
        except Exception as e:
            _logger.error(f"Error al invalidar token para el registro {self.id}: {e}")
            return False

    @api.model
    def limpiar_tokens_antiguos(self):
        """Limpia tokens de cambio de ubicación más antiguos de 24 horas"""
        try:
            from datetime import datetime, timedelta
            yesterday = datetime.now() - timedelta(hours=24)
            
            registros_antiguos = self.search([
                ('location_change_token', '!=', False),
                ('write_date', '<', yesterday)
            ])
            
            if registros_antiguos:
                registros_antiguos.write({'location_change_token': False})
                _logger.info(f"Se limpiaron {len(registros_antiguos)} tokens antiguos")
                
        except Exception as e:
            _logger.error(f"Error al limpiar tokens antiguos: {e}")
    def _notify_vendedora(self):
        return {
            'warning': {
                'title': "Notificación",
                'message': "Estimada vendedora, ya se está notificando a transporte que traigan el equipo.",
                'type': 'notification'
            }
        }
    @api.model
    def cron_evaluador_diario(self):
        _logger.debug("Iniciando cron_evaluador_diario")
        self.evaluar_registros_diarios()

    def evaluar_registros_diarios(self):
        """Evalúa diariamente si hay que traer máquinas del segundo local o Covida."""
        registros_primer_piso = self.search([
            ('ubicacion_id', '=', 'primer_piso'), 
            ('estado_ventas_id', '=', 'sin_revisar')
        ])
        registros_tercer_piso = self.search([
            ('ubicacion_id', '=', 'tercer_piso'), 
            ('estado_ventas_id', '=', 'sin_revisar')
        ])

        if not registros_primer_piso and not registros_tercer_piso:
            registros_a_traer = self.search([
                ('ubicacion_id', 'in', ['segundo_local', 'covida']),
                ('estado_ventas_id', '=', 'sin_revisar')
            ], limit=8)
            
            _logger.info(f"Cron diario: Se encontraron {len(registros_a_traer)} máquinas para traer")
            
            if registros_a_traer:
                transportista_numeros = ['51924894872']
                exitos = 0
                fallos = 0
                
                for registro in registros_a_traer:
                    try:
                        # Obtener el display name de la ubicación
                        ubicacion_display = dict(registro._fields['ubicacion_id'].selection).get(
                            registro.ubicacion_id, 
                            registro.ubicacion_id
                        )
                        
                        mensaje = f"""*Solicitud de traslado de máquina*

    📦 *Detalles del equipo:*
    - Modelo: *{registro.name.name}*
    - Serie: *{registro.serie_id}*
    - Ubicación actual: *{ubicacion_display}*

    📍 *Destino:* Primer piso

    Para registrar el cambio de ubicación cuando llegue la máquina, 
    haga clic en el siguiente enlace:
    {self.crear_url_cambio_ubicacion(registro)}"""

                        _logger.info(f"Enviando mensaje para la máquina {registro.name.name} con serie {registro.serie_id}")

                        # Enviar a cada transportista
                        mensaje_enviado = False
                        for numero in transportista_numeros:
                            try:
                                # 🔥 CAMBIO: Usar API BOOT
                                resultado = registro._send_whatsapp_message_boot(numero, mensaje)
                                
                                if resultado:
                                    _logger.info(f"✅ Mensaje enviado exitosamente a {numero} para máquina {registro.serie_id}")
                                    mensaje_enviado = True
                                else:
                                    _logger.warning(f"⚠️ Fallo al enviar mensaje a {numero} para máquina {registro.serie_id}")
                                    
                            except Exception as e:
                                _logger.error(f"❌ Error al enviar mensaje a {numero} para máquina {registro.serie_id}: {e}")
                        
                        if mensaje_enviado:
                            exitos += 1
                            # Registrar en el chatter del equipo
                            registro.message_post(
                                body=f"✅ Notificación enviada a transportistas en evaluación diaria.",
                                message_type='notification',
                                subtype_xmlid='mail.mt_note'
                            )
                        else:
                            fallos += 1
                            # Registrar fallo en el chatter
                            registro.message_post(
                                body=f"⚠️ No se pudo enviar notificación a transportistas en evaluación diaria.",
                                message_type='notification',
                                subtype_xmlid='mail.mt_note'
                            )
                            
                    except Exception as e:
                        fallos += 1
                        _logger.error(f"❌ Error procesando máquina {registro.serie_id}: {e}", exc_info=True)
                        try:
                            registro.message_post(
                                body=f"❌ Error al procesar notificación de transporte: {str(e)}",
                                message_type='notification',
                                subtype_xmlid='mail.mt_note'
                            )
                        except:
                            pass
                
                _logger.info(
                    f"Cron diario completado: {exitos} éxitos, {fallos} fallos "
                    f"de {len(registros_a_traer)} máquinas procesadas"
                )
            else:
                _logger.info("Cron diario: No hay máquinas para traer")
        else:
            _logger.info(
                f"Cron diario: Hay stock suficiente "
                f"(Primer piso: {len(registros_primer_piso)}, Tercer piso: {len(registros_tercer_piso)})"
            )
    def action_crear_reparaciones(self):
        """ Crea una reparación para cada registro seleccionado en el modelo 'sat.sat'. """
        
        reparacion_model = self.env['reparaciones.reparaciones']
        
        for record in self:
            # Crear la reparación para cada registro seleccionado
            reparacion_model.create({
                'maquina_id': record.id,  # Relaciona la reparación con el registro actual de 'sat.sat'
                # Agrega otros campos necesarios en el modelo 'reparaciones.reparaciones'
            })
            _logger.info(f"Reparación creada para la máquina {record.name.name} con serie {record.serie_id}")
        
        # Mostrar una notificación de éxito
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Reparaciones creadas'),
                'message': _('Se han creado las reparaciones para las máquinas seleccionadas.'),
                'type': 'success',
                'sticky': False,
            }
        }

    def _get_main_reparacion(self):
        """Devuelve la reparación principal para esta máquina.
           Prioriza `reparacion_id` y si no, la última reparación creada."""
        self.ensure_one()
        reparacion = self.reparacion_id
        if not reparacion and self.reparaciones_ids:
            # Tomar la última reparación (por create_date)
            reparacion = self.reparaciones_ids.sorted(lambda r: r.create_date or r.id, reverse=True)[0]
        return reparacion

    def action_print_reparacion_pdf(self):
        """Imprimir / descargar el PDF de la reparación desde la máquina."""
        self.ensure_one()
        reparacion = self._get_main_reparacion()
        if not reparacion:
            raise UserError(_("Esta máquina no tiene ninguna reparación registrada."))

        try:
            report = self.env.ref('sat.action_report_reparaciones_ventas')
        except ValueError:
            raise UserError(_("No se encontró la acción de reporte 'sat.action_report_reparaciones_ventas'."))

        # Imprime el PDF de esa reparación
        return report.report_action(reparacion)

    def action_open_reparacion_gallery(self):
        """Abrir directamente la galería de fotos de la reparación desde la máquina."""
        self.ensure_one()
        reparacion = self._get_main_reparacion()
        if not reparacion:
            raise UserError(_("Esta máquina no tiene ninguna reparación registrada."))

        # Reusa la acción ya definida en reparaciones.reparaciones
        return reparacion.action_open_gallery()


        # ===============================
    #   DASHBOARD LISTA sat.sat
    # ===============================
    @api.model
    def get_sat_dashboard_values(self, domain=False):
        domain = domain or []
        Sat = self.env['sat.sat']

        # IMPORTANTE: Excluir máquinas entregadas del conteo general
        domain_sin_entregadas = domain + [('estado_ventas_id', '!=', 'entregada')]
        
        records = Sat.search(domain_sin_entregadas)
        total_maquinas = len(records)

        # Disponibilidad (sin entregadas)
        total_disponibles = Sat.search_count(domain_sin_entregadas + [
            ('disponibilidad_id', '=', 'disponible')
        ])
        total_separadas = Sat.search_count(domain_sin_entregadas + [
            ('disponibilidad_id', '=', 'separada')
        ])
        total_no_disponibles = Sat.search_count(domain_sin_entregadas + [
            ('disponibilidad_id', '=', 'no_disponible')
        ])

        # Estados de revisión (sin entregadas)
        total_sin_revisar = Sat.search_count(domain_sin_entregadas + [
            ('estado_ventas_id', '=', 'sin_revisar')
        ])
        total_para_revision = Sat.search_count(domain_sin_entregadas + [
            ('estado_ventas_id', '=', 'para_revision')
        ])
        total_en_revision = Sat.search_count(domain_sin_entregadas + [
            ('estado_ventas_id', '=', 'en_revision')
        ])
        total_finalizado = Sat.search_count(domain_sin_entregadas + [
            ('estado_ventas_id', '=', 'finalizado')
        ])
        total_problemas = Sat.search_count(domain_sin_entregadas + [
            ('estado_ventas_id', '=', 'con_problemas')
        ])
        total_de_partes = Sat.search_count(domain_sin_entregadas + [
            ('estado_ventas_id', '=', 'de_partes')
        ])
        
        # Entregadas: solo para mostrar en su propio botón si se necesita
        total_entregada = Sat.search_count(domain + [
            ('estado_ventas_id', '=', 'entregada')
        ])

        today = fields.Date.context_today(self)
        last_7 = today - timedelta(days=7)
        last_30 = today - timedelta(days=30)

        ingresadas_7 = Sat.search_count(domain_sin_entregadas + [
            ('create_date', '>=', last_7)
        ])
        entregadas_30 = Sat.search_count(domain + [
            ('estado_ventas_id', '=', 'entregada'),
            ('fecha_entrega', '>=', last_30)
        ])

        disponibles_records = Sat.search(domain_sin_entregadas + [
            ('disponibilidad_id', '=', 'disponible')
        ])
        stock_value_available = sum(disponibles_records.mapped('precio_compra'))

        precios = [r.precio_compra for r in records if r.precio_compra]
        avg_precio_compra = sum(precios) / len(precios) if precios else 0.0

        currency = self.env.company.currency_id

        return {
            'company_currency_symbol': currency.symbol or '',
            'total_maquinas': total_maquinas,
            'total_disponibles': total_disponibles,
            'total_separadas': total_separadas,
            'total_no_disponibles': total_no_disponibles,
            'total_sin_revisar': total_sin_revisar,
            'total_para_revision': total_para_revision,
            'total_en_revision': total_en_revision,
            'total_finalizado': total_finalizado,
            'total_problemas': total_problemas,
            'total_de_partes': total_de_partes,
            'total_entregada': total_entregada,
            'ingresadas_7': ingresadas_7,
            'entregadas_30': entregadas_30,
            'stock_value_available': round(stock_value_available, 2),
            'avg_precio_compra': round(avg_precio_compra, 2),
        }

    def action_entregar(self):
        """
        Abre el wizard para registrar factura y fecha de entrega.
        """
        self.ensure_one()

        return {
            'name': 'Registrar entrega',
            'type': 'ir.actions.act_window',
            'res_model': 'sat.entrega.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_maquina_id': self.id,
            }
        }


    def action_regresar_a_finalizado(self):
        """
        Regresa el estado de entregada a finalizado
        y limpia factura y fecha.
        """
        for record in self:
            record.write({
                'estado_ventas_id': 'finalizado',
                'factura_venta': False,
                'fecha_entrega': False,
            })

            record.message_post(
                body="Se regresó el estado de 'Entregada' a 'Finalizado'. Se limpiaron factura y fecha de entrega."
            )


    

    
